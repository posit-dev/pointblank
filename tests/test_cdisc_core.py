"""Tests for the CDISC CORE JSON report parser and CORE-backed ConformanceReport.

These tests run entirely against captured CORE report fixtures — no CORE engine dependency. The
fixtures were produced by `core validate -s sdtmig -v 3-4 ... -of JSON` against CORE 0.16.0.
"""

import json
import sys
from pathlib import Path

import pytest

from pointblank.metadata import (
    ConformanceReport,
    CoreFinding,
    CoreRuleResult,
    ParsedCoreReport,
    parse_core_report,
)
from pointblank.metadata._cdisc_core import (
    STATUS_ERROR,
    STATUS_ISSUE,
    STATUS_SKIPPED,
    STATUS_SUCCESS,
    CoreExecutionError,
    CoreNotFoundError,
    _CoreRunner,
    _materialize_datasets,
    _normalize_version,
    _resolve_core_command,
    _write_xpt,
)

_FIXTURES = Path(__file__).parent / "metadata_fixtures" / "cdisc_core"


def _load(name: str) -> dict:
    return json.loads((_FIXTURES / name).read_text())


# A stand-in for the CORE CLI: copies a source JSON to `<stem>.<ext>` and records its argv to
# `<stem>.argv.json`, so tests can drive the runner and assert the built command — no real CORE.
_FAKE_CORE_TEMPLATE = """\
import json, shutil, sys
args = sys.argv[1:]
def opt(flag):
    return args[args.index(flag) + 1] if flag in args else None
stem = opt("-o")
fmt = opt("-of") or "JSON"
ext = {{"JSON": "json", "XLSX": "xlsx", "CSV": "csv"}}[fmt]
with open(stem + ".argv.json", "w") as f:
    json.dump(args, f)
shutil.copy({src!r}, stem + "." + ext)
sys.exit({exit_code})
"""


def _make_fake_core(tmp_path: Path, src_json: Path, exit_code: int = 0) -> Path:
    script = tmp_path / "fakecore.py"
    script.write_text(_FAKE_CORE_TEMPLATE.format(src=str(src_json), exit_code=exit_code))
    return script


def _fake_runner(tmp_path: Path, exit_code: int = 0) -> _CoreRunner:
    src = _FIXTURES / "core_report_trimmed.json"
    script = _make_fake_core(tmp_path, src, exit_code=exit_code)
    return _CoreRunner(core=[sys.executable, str(script)])


def _fake_core_cmd(tmp_path: Path) -> list:
    """A `core=` command list (usable with validate_conformance) backed by the fake CORE."""
    script = _make_fake_core(tmp_path, _FIXTURES / "core_report_trimmed.json")
    return [sys.executable, str(script)]


@pytest.fixture
def full_report() -> dict:
    return _load("core_report_full.json")


@pytest.fixture
def trimmed_report() -> dict:
    return _load("core_report_trimmed.json")


# ── Parser ───────────────────────────────────────────────────────────────────


def test_parse_full_report_provenance(full_report):
    parsed = parse_core_report(full_report)
    assert isinstance(parsed, ParsedCoreReport)
    assert parsed.standard == "SDTMIG"
    assert parsed.version == "V3.4"
    assert parsed.engine_version == "0.16.0"
    assert len(parsed.rules) == 430
    assert len(parsed.datasets) == 1


def test_parse_status_counts(full_report):
    parsed = parse_core_report(full_report)
    counts = parsed.status_counts()
    assert counts == {
        STATUS_SKIPPED: 344,
        STATUS_SUCCESS: 78,
        STATUS_ISSUE: 6,
        STATUS_ERROR: 2,
    }


def test_parse_findings_typed(full_report):
    parsed = parse_core_report(full_report)
    assert all(isinstance(f, CoreFinding) for f in parsed.findings)
    f = parsed.findings[0]
    assert f.rule_id == "CORE-000357"
    assert f.dataset == "TEST_DATASET"
    assert f.row == 1
    assert f.variables == ["dataset_name"]
    assert f.values == ["TEST_DATASET"]
    # Empty-string sentinels normalized to None
    assert f.usubjid is None
    assert f.seq is None


def test_parse_rules_typed_and_is_failing(full_report):
    parsed = parse_core_report(full_report)
    assert all(isinstance(r, CoreRuleResult) for r in parsed.rules)
    failing = parsed.failing_rules()
    # 6 ISSUE REPORTED + 2 EXECUTION ERROR = 8 failing rules
    assert len(failing) == 8
    assert all(r.is_failing for r in failing)
    success = [r for r in parsed.rules if r.status == STATUS_SUCCESS]
    assert success and not success[0].is_failing


def test_parse_total_issues_and_all_passed(full_report):
    parsed = parse_core_report(full_report)
    assert parsed.n_total_issues == sum(s.issues for s in parsed.issue_summary)
    assert parsed.all_passed is False


def test_parse_accepts_already_parsed_via_report(trimmed_report):
    # from_core_report should accept both a dict and a ParsedCoreReport
    parsed = parse_core_report(trimmed_report)
    rep = ConformanceReport.from_core_report(parsed)
    assert rep.is_core
    assert rep.core is parsed


def test_all_passed_true_when_no_failing_rules():
    report = {
        "Conformance_Details": {"Standard": "SDTMIG", "Version": "V3.4"},
        "Dataset_Details": [],
        "Issue_Summary": [],
        "Issue_Details": [],
        "Rules_Report": [
            {"core_id": "CORE-1", "status": "SUCCESS"},
            {"core_id": "CORE-2", "status": "SKIPPED"},
        ],
    }
    parsed = parse_core_report(report)
    assert parsed.all_passed is True


def test_all_passed_falls_back_to_issue_summary_when_no_rules():
    report = {
        "Conformance_Details": {},
        "Issue_Summary": [{"dataset": "DM", "core_id": "CORE-1", "message": "x", "issues": 3}],
        "Issue_Details": [],
        "Rules_Report": [],
    }
    parsed = parse_core_report(report)
    assert parsed.all_passed is False


def test_parse_raises_on_non_dict():
    with pytest.raises(TypeError):
        parse_core_report(["not", "a", "dict"])


def test_parse_raises_on_unrecognized_dict():
    with pytest.raises(ValueError):
        parse_core_report({"foo": "bar", "baz": 1})


def test_parse_raw_report_ignores_results_data(full_report):
    # The --raw-report variant adds a results_data key; parser should ignore it.
    raw = dict(full_report)
    raw["results_data"] = [{"anything": 1}]
    parsed = parse_core_report(raw)
    assert len(parsed.rules) == 430


def test_as_int_coercion_edge_cases():
    report = {
        "Conformance_Details": {},
        "Issue_Summary": [{"dataset": "DM", "core_id": "C", "message": "m", "issues": ""}],
        "Issue_Details": [{"core_id": "C", "message": "m", "dataset": "DM", "row": "5"}],
        "Rules_Report": [],
    }
    parsed = parse_core_report(report)
    assert parsed.issue_summary[0].issues == 0  # "" -> 0
    assert parsed.findings[0].row == 5  # "5" -> 5


# ── ConformanceReport (CORE form) ──────────────────────────────────────────────


def test_report_is_core_and_all_passed(full_report):
    rep = ConformanceReport.from_core_report(full_report, agency="FDA")
    assert rep.is_core is True
    assert rep.agency == "FDA"
    assert rep.all_passed() is False
    assert rep.n_datasets == 1


def test_report_summary_core(full_report):
    rep = ConformanceReport.from_core_report(full_report)
    s = rep.summary()
    assert s["standard"] == "SDTMIG"
    assert s["version"] == "V3.4"
    assert s["engine_version"] == "0.16.0"
    assert s["n_rules"] == 430
    assert s["n_issues"] == 8
    assert s["all_passed"] is False
    assert s["status_counts"][STATUS_ISSUE] == 6


def test_report_issues_core(full_report):
    rep = ConformanceReport.from_core_report(full_report)
    issues = rep.issues()
    assert len(issues) == 8
    assert all({"dataset", "rule_id", "message", "issues", "status"} <= set(i) for i in issues)
    # every issue-summary rule id resolves to a status in the (consistent) full fixture
    assert all(i["status"] is not None for i in issues)


def test_report_issues_status_filter(full_report):
    rep = ConformanceReport.from_core_report(full_report)
    errs = rep.issues(status=STATUS_ERROR)
    assert {i["rule_id"] for i in errs} == {"CORE-000929", "CORE-001081"}
    assert all(i["status"] == STATUS_ERROR for i in errs)


def test_report_findings_and_rules_accessors(full_report):
    rep = ConformanceReport.from_core_report(full_report)
    findings = rep.findings()
    assert findings and isinstance(findings[0], CoreFinding)
    all_rules = rep.rules()
    assert len(all_rules) == 430
    issue_rules = rep.rules(status=STATUS_ISSUE)
    assert len(issue_rules) == 6
    assert all(r.status == STATUS_ISSUE for r in issue_rules)


def test_report_repr_and_html_core(full_report):
    rep = ConformanceReport.from_core_report(full_report, agency="FDA")
    text = repr(rep)
    assert "ConformanceReport (CORE)" in text
    assert "SDTMIG V3.4" in text
    assert "FAIL" in text
    html = rep._repr_html_()
    assert "CORE" in html
    assert "SDTMIG" in html
    assert "CORE-000357" in html  # an issue rule id appears in the table


def test_trimmed_fixture_is_internally_consistent(trimmed_report):
    # Every rule referenced by Issue_Summary must resolve to a status (no None) in the trimmed set.
    rep = ConformanceReport.from_core_report(trimmed_report)
    assert all(i["status"] is not None for i in rep.issues())


# ── Native/CORE separation ─────────────────────────────────────────────────────


def test_native_rules_report_properties():
    # A native rules-engine report: is_core False, is_rules True, has rule results.
    import pandas as pd

    import pointblank as pb
    from pointblank.metadata._conformance.result import NativeRuleResult

    dm = pd.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["S1-001"],
            "SUBJID": ["001"],
            "ARMCD": ["A"],
            "ARM": ["Arm A"],
            "COUNTRY": ["USA"],
        }
    )
    rep = pb.SubmissionPackage(datasets={"DM": dm}).validate_conformance()
    assert rep.is_core is False
    assert rep.is_rules is True
    # findings() returns a list (may be empty for clean data)
    assert isinstance(rep.findings(), list)
    # rules() returns NativeRuleResult objects
    rules = rep.rules()
    assert len(rules) > 0
    assert all(isinstance(r, NativeRuleResult) for r in rules)


# ── Dataset materialization (_write_xpt / _materialize_datasets) ────────────────


def _dm_df():
    import pandas as pd

    return pd.DataFrame(
        {
            "STUDYID": ["S1", "S1"],
            "DOMAIN": ["DM", "DM"],
            "USUBJID": ["S1-001", "S1-002"],
        }
    )


def test_write_xpt_roundtrip(tmp_path):
    pyreadstat = pytest.importorskip("pyreadstat")
    df = _dm_df()
    out = _write_xpt(df, tmp_path / "dm.xpt", table_name="DM", file_label="Demographics")
    assert out.exists()
    back, meta = pyreadstat.read_xport(str(out))
    assert list(back.columns) == ["STUDYID", "DOMAIN", "USUBJID"]
    assert len(back) == 2
    assert meta.table_name == "DM"


def test_write_xpt_truncates_member_name(tmp_path):
    pyreadstat = pytest.importorskip("pyreadstat")
    df = _dm_df()
    out = _write_xpt(df, tmp_path / "x.xpt", table_name="SUPPLONGNAME")
    _back, meta = pyreadstat.read_xport(str(out))
    assert len(meta.table_name) <= 8
    assert meta.table_name == "SUPPLONG"


def test_write_xpt_accepts_polars(tmp_path):
    pytest.importorskip("pyreadstat")
    pl = pytest.importorskip("polars")
    df = pl.from_pandas(_dm_df())
    out = _write_xpt(df, tmp_path / "dm.xpt", table_name="DM")
    assert out.exists()


def test_materialize_datasets(tmp_path):
    pytest.importorskip("pyreadstat")
    written = _materialize_datasets({"DM": _dm_df(), "AE": _dm_df()}, tmp_path / "mat")
    assert set(written) == {"DM", "AE"}
    assert written["DM"].name == "dm.xpt"
    assert written["AE"].name == "ae.xpt"
    assert all(p.exists() for p in written.values())


# ── CORE command resolution / discovery ────────────────────────────────────────


def test_normalize_version():
    assert _normalize_version("3.4") == "3-4"
    assert _normalize_version("3-4") == "3-4"
    assert _normalize_version("1.1") == "1-1"


def test_resolve_core_command_explicit_str():
    assert _resolve_core_command("core") == ["core"]


def test_resolve_core_command_explicit_sequence():
    assert _resolve_core_command(["python", "core.py"]) == ["python", "core.py"]


def test_resolve_core_command_from_env(monkeypatch):
    monkeypatch.setenv("POINTBLANK_CDISC_CORE", "python /path/core.py")
    assert _resolve_core_command(None) == ["python", "/path/core.py"]


def test_resolve_core_command_not_found(monkeypatch):
    monkeypatch.delenv("POINTBLANK_CDISC_CORE", raising=False)
    # Force PATH discovery to fail
    monkeypatch.setattr("pointblank.metadata._cdisc_core.shutil.which", lambda name: None)
    with pytest.raises(CoreNotFoundError):
        _resolve_core_command(None)


# ── CORE runner (driven by a fake CORE script) ─────────────────────────────────


def test_runner_command_property(tmp_path):
    runner = _fake_runner(tmp_path)
    assert runner.command[0] == sys.executable
    assert runner.command[1].endswith("fakecore.py")


def test_run_validate_produces_and_parses(tmp_path):
    runner = _fake_runner(tmp_path)
    parsed = runner.validate_to_report(
        data_dir=tmp_path,
        standard="sdtmig",
        version="3.4",
        output_stem=tmp_path / "out",
    )
    assert isinstance(parsed, ParsedCoreReport)
    assert parsed.standard == "SDTMIG"
    assert len(parsed.rules) == 12


def test_run_validate_builds_expected_command(tmp_path):
    runner = _fake_runner(tmp_path)
    define = tmp_path / "define.xml"
    define.write_text("<define/>")
    runner.run_validate(
        data_dir=tmp_path / "data",
        standard="sdtmig",
        version="3.4",
        output_stem=tmp_path / "out",
        define_xml=define,
        controlled_terminology=["sdtmct-2024-03-29", "sdtmct-2023-12-15"],
        cache=tmp_path / "cache",
        raw_report=True,
    )
    argv = json.loads((tmp_path / "out.argv.json").read_text())
    # Positional/flag checks
    assert argv[0] == "validate"
    assert "-s" in argv and argv[argv.index("-s") + 1] == "sdtmig"
    assert argv[argv.index("-v") + 1] == "3-4"  # hyphenated
    assert argv[argv.index("-of") + 1] == "JSON"
    assert argv[argv.index("-dxp") + 1] == str(define)
    assert argv[argv.index("-ca") + 1] == str(tmp_path / "cache")
    assert argv.count("-ct") == 2
    assert "-rr" in argv


def test_run_validate_single_ct_string(tmp_path):
    runner = _fake_runner(tmp_path)
    runner.run_validate(
        data_dir=tmp_path,
        standard="sdtmig",
        version="3.4",
        output_stem=tmp_path / "out",
        controlled_terminology="sdtmct-2024-03-29",
    )
    argv = json.loads((tmp_path / "out.argv.json").read_text())
    assert argv.count("-ct") == 1
    assert argv[argv.index("-ct") + 1] == "sdtmct-2024-03-29"


def test_run_validate_unsupported_format(tmp_path):
    runner = _fake_runner(tmp_path)
    with pytest.raises(ValueError):
        runner.run_validate(
            data_dir=tmp_path,
            standard="sdtmig",
            version="3.4",
            output_stem=tmp_path / "out",
            output_format="PDF",
        )


def test_run_validate_nonzero_exit_raises(tmp_path):
    runner = _fake_runner(tmp_path, exit_code=2)
    with pytest.raises(CoreExecutionError):
        runner.run_validate(
            data_dir=tmp_path,
            standard="sdtmig",
            version="3.4",
            output_stem=tmp_path / "out",
        )


def test_run_validate_missing_output_raises(tmp_path):
    # A fake core that exits 0 but writes nothing.
    script = tmp_path / "noop.py"
    script.write_text("import sys; sys.exit(0)")
    runner = _CoreRunner(core=[sys.executable, str(script)])
    with pytest.raises(CoreExecutionError):
        runner.run_validate(
            data_dir=tmp_path,
            standard="sdtmig",
            version="3.4",
            output_stem=tmp_path / "out",
        )


def test_runner_passes_cwd(tmp_path):
    # CORE resolves its bundled resources relative to cwd; verify the runner honors cwd=.
    src = _FIXTURES / "core_report_trimmed.json"
    run_dir = tmp_path / "coreroot"
    run_dir.mkdir()
    script = tmp_path / "cwdcore.py"
    script.write_text(
        "import json, os, shutil, sys\n"
        "args = sys.argv[1:]\n"
        "stem = args[args.index('-o') + 1]\n"
        "open(stem + '.cwd.txt', 'w').write(os.getcwd())\n"
        f"shutil.copy({str(src)!r}, stem + '.json')\n"
    )
    runner = _CoreRunner(core=[sys.executable, str(script)], cwd=run_dir)
    assert runner.cwd == str(run_dir)
    runner.run_validate(
        data_dir=tmp_path,
        standard="sdtmig",
        version="3.4",
        output_stem=tmp_path / "out",
    )
    recorded = (tmp_path / "out.cwd.txt").read_text()
    assert Path(recorded).resolve() == run_dir.resolve()


def test_runner_launch_failure_raises(tmp_path):
    runner = _CoreRunner(core=[str(tmp_path / "does-not-exist-binary")])
    with pytest.raises(CoreExecutionError):
        runner.run_validate(
            data_dir=tmp_path,
            standard="sdtmig",
            version="3.4",
            output_stem=tmp_path / "out",
        )


# ── validate_conformance(engine="core") wiring ─────────────────────────────────


def _dm_pkg():
    import pointblank as pb

    return pb.SubmissionPackage(
        datasets={"DM": _dm_df()}, standard="sdtmig", standard_version="3.4"
    )


def test_validate_conformance_core_materializes_in_memory(tmp_path):
    pytest.importorskip("pyreadstat")
    rep = _dm_pkg().validate_conformance(
        engine="core", agency="FDA", core=_fake_core_cmd(tmp_path), workdir=tmp_path / "w"
    )
    assert rep.is_core is True
    assert rep.agency == "FDA"
    assert rep.summary()["standard"] == "SDTMIG"
    # in-memory datasets were materialized to XPT for CORE
    assert (tmp_path / "w" / "data" / "dm.xpt").exists()


def test_validate_conformance_core_version_hyphenated(tmp_path):
    pytest.importorskip("pyreadstat")
    _dm_pkg().validate_conformance(
        engine="core", core=_fake_core_cmd(tmp_path), workdir=tmp_path / "w"
    )
    argv = json.loads((tmp_path / "w" / "core_report.argv.json").read_text())
    assert argv[argv.index("-v") + 1] == "3-4"
    assert argv[argv.index("-s") + 1] == "sdtmig"


def test_validate_conformance_core_folder_passthrough(tmp_path):
    import shutil

    import pointblank as pb

    folder = tmp_path / "study"
    folder.mkdir()
    shutil.copy(pb.load_metadata_example("dm.xpt"), folder / "dm.xpt")
    shutil.copy(pb.load_metadata_example("define.xml"), folder / "define.xml")

    pkg = pb.SubmissionPackage.from_folder(folder)
    assert pkg.source_folder == str(folder)

    rep = pkg.validate_conformance(
        engine="core", core=_fake_core_cmd(tmp_path), workdir=tmp_path / "w"
    )
    assert rep.is_core
    argv = json.loads((tmp_path / "w" / "core_report.argv.json").read_text())
    # CORE was pointed at the source folder directly, with define via -dxp
    assert argv[argv.index("-d") + 1] == str(folder)
    assert "-dxp" in argv


def test_validate_conformance_rejects_bad_engine():
    with pytest.raises(ValueError):
        _dm_pkg().validate_conformance(engine="bogus")


def test_validate_conformance_native_default_unchanged():
    # engine defaults to native and yields a native report
    rep = _dm_pkg().validate_conformance()
    assert rep.is_core is False


# ── validate_cdisc_submission() convenience entry point ─────────────────────────


def test_validate_cdisc_submission_from_dict(tmp_path):
    pytest.importorskip("pyreadstat")
    import pointblank as pb

    rep = pb.validate_cdisc_submission(
        {"DM": _dm_df()},
        standard="sdtmig",
        version="3.4",
        core=_fake_core_cmd(tmp_path),
        workdir=tmp_path / "w",
    )
    assert rep.is_core
    assert rep.summary()["n_rules"] == 12


def test_validate_cdisc_submission_from_folder(tmp_path):
    import shutil

    import pointblank as pb

    folder = tmp_path / "study"
    folder.mkdir()
    shutil.copy(pb.load_metadata_example("dm.xpt"), folder / "dm.xpt")

    rep = pb.validate_cdisc_submission(
        folder, core=_fake_core_cmd(tmp_path), workdir=tmp_path / "w"
    )
    assert rep.is_core


def test_validate_cdisc_submission_from_package(tmp_path):
    pytest.importorskip("pyreadstat")
    import pointblank as pb

    pkg = _dm_pkg()
    rep = pb.validate_cdisc_submission(pkg, core=_fake_core_cmd(tmp_path), workdir=tmp_path / "w")
    assert rep.is_core
    assert rep.package is pkg


def test_validate_cdisc_submission_bad_type():
    import pointblank as pb

    with pytest.raises(TypeError):
        pb.validate_cdisc_submission(12345)


def test_validate_cdisc_submission_nonexistent_folder():
    import pointblank as pb

    with pytest.raises(NotADirectoryError):
        pb.validate_cdisc_submission("/no/such/folder/here")


def test_validate_cdisc_submission_exported():
    import pointblank as pb

    assert pb.validate_cdisc_submission is not None
    assert "validate_cdisc_submission" in pb.__all__


# ── ConformanceReport export methods ──────────────────────────────────────────


def test_to_json_core_structure(tmp_path, full_report):
    rep = ConformanceReport.from_core_report(full_report)
    dest = rep.to_json(tmp_path / "report.json")
    assert dest.exists()
    import json

    data = json.loads(dest.read_text())
    assert "Conformance_Details" in data
    assert "Issue_Summary" in data
    assert "Issue_Details" in data
    assert "Rules_Report" in data
    assert len(data["Rules_Report"]) == 430
    assert all("core_id" in r for r in data["Rules_Report"])


def test_to_json_core_roundtrips_through_parser(tmp_path, full_report):
    rep = ConformanceReport.from_core_report(full_report)
    dest = rep.to_json(tmp_path / "report.json")
    import json

    from pointblank.metadata import parse_core_report

    reparsed = parse_core_report(json.loads(dest.read_text()))
    assert reparsed.standard == rep.core.standard
    assert len(reparsed.rules) == len(rep.core.rules)
    assert reparsed.n_total_issues == rep.core.n_total_issues


def test_to_json_core_creates_parent_dirs(tmp_path, trimmed_report):
    rep = ConformanceReport.from_core_report(trimmed_report)
    dest = rep.to_json(tmp_path / "nested" / "dir" / "report.json")
    assert dest.exists()


def test_to_json_native(tmp_path):
    import json

    import pandas as pd

    import pointblank as pb

    dm = pd.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["S1-001"],
            "SUBJID": ["001"],
            "ARMCD": ["A"],
            "ARM": ["A"],
            "COUNTRY": ["USA"],
        }
    )
    rep = pb.SubmissionPackage(datasets={"DM": dm}).validate_conformance()
    dest = rep.to_json(tmp_path / "native_report.json")
    assert dest.exists()
    data = json.loads(dest.read_text())
    # Rules-engine native report has flat summary and issues list.
    assert "summary" in data
    assert "issues" in data
    s = data["summary"]
    assert s["engine"] == "built-in"
    assert "standard" in s
    assert "n_rules" in s


def test_to_excel_core_sheets(tmp_path, full_report):
    pytest.importorskip("openpyxl")
    import openpyxl

    rep = ConformanceReport.from_core_report(full_report)
    dest = rep.to_excel(tmp_path / "report.xlsx")
    assert dest.exists()
    wb = openpyxl.load_workbook(dest)
    assert "Issue_Summary" in wb.sheetnames
    assert "Issue_Details" in wb.sheetnames
    assert "Rules_Report" in wb.sheetnames
    assert "Conformance_Details" in wb.sheetnames


def test_to_excel_core_row_counts(tmp_path, full_report):
    pytest.importorskip("openpyxl")
    import openpyxl

    rep = ConformanceReport.from_core_report(full_report)
    dest = rep.to_excel(tmp_path / "report.xlsx")
    wb = openpyxl.load_workbook(dest)
    # Rules_Report sheet: 1 header row + 430 data rows
    rules_rows = wb["Rules_Report"].max_row
    assert rules_rows == 431


def test_to_excel_native(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    import pandas as pd

    import pointblank as pb

    dm = pd.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["S1-001"],
            "SUBJID": ["001"],
            "ARMCD": ["A"],
            "ARM": ["A"],
            "COUNTRY": ["USA"],
        }
    )
    rep = pb.SubmissionPackage(datasets={"DM": dm}).validate_conformance()
    dest = rep.to_excel(tmp_path / "native_report.xlsx")
    assert dest.exists()
    wb = openpyxl.load_workbook(dest)
    # Rules-engine native report produces Rules_Report and Summary sheets.
    assert "Rules_Report" in wb.sheetnames
    assert "Summary" in wb.sheetnames


def test_to_excel_missing_openpyxl(tmp_path, full_report, monkeypatch):
    import builtins

    real_import = builtins.__import__

    def mock_import(name, *args, **kwargs):
        if name == "openpyxl":
            raise ImportError("mocked missing openpyxl")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", mock_import)
    rep = ConformanceReport.from_core_report(full_report)
    with pytest.raises(ImportError, match="openpyxl"):
        rep.to_excel(tmp_path / "report.xlsx")
