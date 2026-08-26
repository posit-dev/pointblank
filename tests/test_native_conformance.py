"""Tests for the native CDISC conformance rule engine."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import polars as pl
import pytest

import pointblank as pb
from pointblank.metadata._conformance import (
    ControlledTerminology,
    NativeConformanceEngine,
    NativeConformanceResult,
    NativeRowFinding,
    NativeRuleResult,
    RuleLoader,
)
from pointblank.metadata._conformance.evaluator import evaluate_conditions, is_iso8601
from pointblank.metadata._conformance.operations import apply_operations
from pointblank.metadata._conformance.result import STATUS_FAIL, STATUS_PASS
from pointblank.metadata._types import MetadataImport, MetadataPackage, VariableMetadata
import narwhals as nw


# ── Fixtures ──────────────────────────────────────────────────────────────────


def _clean_dm(backend="polars"):
    data = {
        "STUDYID": ["S001", "S001"],
        "DOMAIN": ["DM", "DM"],
        "USUBJID": ["S001-001", "S001-002"],
        "SUBJID": ["001", "002"],
        "RFSTDTC": ["2024-01-01", "2024-01-02"],
        "RFENDTC": ["2024-06-30", "2024-06-30"],
        "SITEID": ["001", "001"],
        "AGE": [45.0, 32.0],
        "AGEU": ["YEARS", "YEARS"],
        "SEX": ["M", "F"],
        "RACE": ["WHITE", "ASIAN"],
        "ETHNIC": ["NOT HISPANIC OR LATINO", "NOT HISPANIC OR LATINO"],
        "COUNTRY": ["USA", "GBR"],
        "ARMCD": ["A", "B"],
        "ARM": ["Arm A", "Arm B"],
        "ACTARMCD": ["A", "B"],
        "ACTARM": ["Arm A", "Arm B"],
        "DMDTC": ["2024-01-01", "2024-01-02"],
    }
    if backend == "pandas":
        return pd.DataFrame(data)
    return pl.DataFrame(data)


@pytest.fixture
def ct():
    return ControlledTerminology.load_default()


@pytest.fixture
def engine():
    return NativeConformanceEngine("sdtmig", "3.4")


def _clean_ta() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "STUDYID": ["S001", "S001"],
            "DOMAIN": ["TA", "TA"],
            "ARMCD": ["A", "B"],
            "ARM": ["Arm A", "Arm B"],
            "TAETORD": [1, 1],
            "EPOCH": ["TREATMENT", "TREATMENT"],
            "ELEMENT": ["Element 1", "Element 1"],
            "ETCD": ["ET1", "ET1"],
            "TASEQ": [1, 2],
        }
    )


def _clean_ts() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "STUDYID": ["S001"],
            "DOMAIN": ["TS"],
            "TSSEQ": [1],
            "TSPARMCD": ["PLANSUB"],
            "TSPARM": ["Planned Number of Subjects"],
            "TSVAL": ["100"],
        }
    )


@pytest.fixture
def clean_result(engine):
    return engine.run({"DM": _clean_dm(), "TA": _clean_ta(), "TS": _clean_ts()})


# ── RuleLoader ────────────────────────────────────────────────────────────────


def test_rule_loader_loads_sdtmig_34():
    rules = RuleLoader.load("sdtmig", "3.4")
    assert len(rules) >= 20


def test_rule_loader_rule_types_coverage():
    rules = RuleLoader.load("sdtmig", "3.4")
    types = {r.rule_type for r in rules}
    assert "RECORD_CHECK" in types
    assert "DATASET_CONTENTS_CHECK" in types
    assert "DATASET_METADATA_CHECK" in types
    assert "DOMAIN_PRESENCE_CHECK" in types


def test_rule_loader_available():
    available = RuleLoader.available()
    assert ("sdtmig", "3.4") in available


def test_rule_loader_missing_raises():
    with pytest.raises(FileNotFoundError, match="No bundled rule catalog"):
        RuleLoader.load("imaginary", "99.0")


def test_rule_loader_filter_by_type():
    rules = RuleLoader.load("sdtmig", "3.4", rule_types=["RECORD_CHECK"])
    assert all(r.rule_type == "RECORD_CHECK" for r in rules)


def test_rule_loader_catalog_metadata():
    meta = RuleLoader.catalog_metadata("sdtmig", "3.4")
    assert "standard" in meta
    assert meta["standard"] == "sdtmig"
    assert "checksum" in meta


# ── ControlledTerminology ─────────────────────────────────────────────────────


def test_ct_load_default(ct):
    assert len(ct.packages) == 1
    assert "SEX" in ct


def test_ct_valid_term(ct):
    assert ct.is_valid("SEX", "M")
    assert ct.is_valid("SEX", "F")


def test_ct_invalid_term(ct):
    assert not ct.is_valid("SEX", "Q")


def test_ct_null_passes(ct):
    # Null values pass by convention (separate not-null rule handles them).
    assert ct.is_valid("SEX", None)


def test_ct_unknown_codelist_passes(ct):
    assert ct.is_valid("NONEXISTENT_CODELIST", "ANYTHING")


def test_ct_available():
    available = ControlledTerminology.available()
    assert len(available) >= 1


def test_ct_missing_package_raises():
    with pytest.raises(FileNotFoundError):
        ControlledTerminology.load(["no-such-package-2099-01-01"])


def test_ct_route_codelist(ct):
    assert ct.is_valid("ROUTE", "ORAL")
    assert ct.is_valid("ROUTE", "INTRAVENOUS")
    assert not ct.is_valid("ROUTE", "PURPLE")


def test_ct_frm_codelist(ct):
    assert ct.is_valid("FRM", "TABLET")
    assert ct.is_valid("FRM", "CAPSULE")
    assert not ct.is_valid("FRM", "MYSTERY")


def test_ct_epoch_codelist(ct):
    assert ct.is_valid("EPOCH", "SCREENING")
    assert ct.is_valid("EPOCH", "TREATMENT")
    assert not ct.is_valid("EPOCH", "PHASE 99")


def test_ct_freq_codelist(ct):
    assert ct.is_valid("FREQ", "QD")
    assert ct.is_valid("FREQ", "BID")
    assert not ct.is_valid("FREQ", "WHENEVER")


# ── Evaluator ─────────────────────────────────────────────────────────────────


def _nw_df(data: dict) -> nw.DataFrame:
    return nw.from_native(pl.DataFrame(data), eager_only=True)


def test_evaluator_is_null():
    df = _nw_df({"x": [1, None, 3]})
    mask = evaluate_conditions(df, {"all": [{"name": "x", "operator": "is_null", "value": None}]})
    assert mask.to_list() == [False, True, False]


def test_evaluator_is_not_null():
    df = _nw_df({"x": [1, None, 3]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "is_not_null", "value": None}]}
    )
    assert mask.to_list() == [True, False, True]


def test_evaluator_equal_to():
    df = _nw_df({"x": [1, 2, 3]})
    mask = evaluate_conditions(df, {"all": [{"name": "x", "operator": "equal_to", "value": 2}]})
    assert mask.to_list() == [False, True, False]


def test_evaluator_not_equal_to():
    df = _nw_df({"x": [1, 2, 3]})
    mask = evaluate_conditions(df, {"all": [{"name": "x", "operator": "not_equal_to", "value": 2}]})
    assert mask.to_list() == [True, False, True]


def test_evaluator_greater_than():
    df = _nw_df({"x": [1, 5, 3]})
    mask = evaluate_conditions(df, {"all": [{"name": "x", "operator": "greater_than", "value": 3}]})
    assert mask.to_list() == [False, True, False]


def test_evaluator_is_in():
    df = _nw_df({"x": ["A", "B", "C"]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "is_in", "value": ["A", "C"]}]}
    )
    assert mask.to_list() == [True, False, True]


def test_evaluator_not_in():
    df = _nw_df({"x": ["A", "B", "C"]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "not_in", "value": ["A", "C"]}]}
    )
    assert mask.to_list() == [False, True, False]


def test_evaluator_contains():
    df = _nw_df({"x": ["hello world", "foo", "world"]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "contains", "value": "world"}]}
    )
    assert mask.to_list() == [True, False, True]


def test_evaluator_any_combinator():
    df = _nw_df({"x": [1, 2, 3], "y": [10, 20, 30]})
    mask = evaluate_conditions(
        df,
        {
            "any": [
                {"name": "x", "operator": "equal_to", "value": 1},
                {"name": "y", "operator": "equal_to", "value": 30},
            ]
        },
    )
    assert mask.to_list() == [True, False, True]


def test_evaluator_not_combinator():
    df = _nw_df({"x": [1, 2, 3]})
    mask = evaluate_conditions(df, {"not": {"name": "x", "operator": "equal_to", "value": 2}})
    assert mask.to_list() == [True, False, True]


def test_evaluator_empty_conditions_returns_all_false():
    df = _nw_df({"x": [1, 2, 3]})
    mask = evaluate_conditions(df, {})
    assert all(not v for v in mask.to_list())


def test_evaluator_equal_to_column():
    df = _nw_df({"a": [1, 2, 3], "b": [1, 99, 3]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "a", "operator": "equal_to_column", "value": "b"}]}
    )
    assert mask.to_list() == [True, False, True]


def test_evaluator_greater_than_or_equal_to():
    df = _nw_df({"x": [1, 3, 5]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "greater_than_or_equal_to", "value": 3}]}
    )
    assert mask.to_list() == [False, True, True]


def test_evaluator_less_than_or_equal_to():
    df = _nw_df({"x": [1, 3, 5]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "less_than_or_equal_to", "value": 3}]}
    )
    assert mask.to_list() == [True, True, False]


def test_evaluator_not_contains():
    df = _nw_df({"x": ["hello world", "foo", "world"]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "not_contains", "value": "world"}]}
    )
    assert mask.to_list() == [False, True, False]


def test_evaluator_starts_with():
    df = _nw_df({"x": ["hello", "world", "hi"]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "starts_with", "value": "h"}]}
    )
    assert mask.to_list() == [True, False, True]


def test_evaluator_ends_with():
    df = _nw_df({"x": ["hello", "world", "told"]})
    mask = evaluate_conditions(df, {"all": [{"name": "x", "operator": "ends_with", "value": "ld"}]})
    assert mask.to_list() == [False, True, True]


def test_evaluator_matches_regex():
    df = _nw_df({"x": ["abc123", "hello", "X99"]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "x", "operator": "matches_regex", "value": r"\d+"}]}
    )
    assert mask.to_list() == [True, False, True]


def test_evaluator_not_equal_to_column():
    df = _nw_df({"a": [1, 2, 3], "b": [1, 99, 3]})
    mask = evaluate_conditions(
        df, {"all": [{"name": "a", "operator": "not_equal_to_column", "value": "b"}]}
    )
    assert mask.to_list() == [False, True, False]


def test_evaluator_unknown_operator_raises():
    from pointblank.metadata._conformance.evaluator import _compile_leaf

    with pytest.raises(ValueError, match="Unknown operator"):
        _compile_leaf({"name": "x", "operator": "does_not_exist", "value": None})


def test_iso8601_valid():
    assert is_iso8601("2024-01-15")
    assert is_iso8601("2024-01")
    assert is_iso8601("2024")
    assert is_iso8601("2024-01-15T10:30:00")
    assert is_iso8601("2024-01-15T10:30:00Z")
    assert is_iso8601("2024-01-15T10:30:00+05:30")


def test_iso8601_invalid():
    assert not is_iso8601("01/15/2024")
    assert not is_iso8601("not-a-date")
    assert not is_iso8601("")
    assert not is_iso8601(None)


# ── Operations ────────────────────────────────────────────────────────────────


def test_op_codelist_check_valid(ct):
    df = _nw_df({"SEX": ["M", "F", "Q"]})
    result = apply_operations(
        df, [{"operator": "codelist_check", "params": {"column": "SEX", "codelist": "SEX"}}], ct, {}
    )
    assert "_pb_SEX_valid" in result.columns
    assert result["_pb_SEX_valid"].to_list() == [True, True, False]


def test_op_codelist_check_null_passes(ct):
    df = _nw_df({"SEX": ["M", None, "F"]})
    result = apply_operations(
        df, [{"operator": "codelist_check", "params": {"column": "SEX", "codelist": "SEX"}}], ct, {}
    )
    assert result["_pb_SEX_valid"].to_list() == [True, True, True]


def test_op_codelist_check_missing_column(ct):
    df = _nw_df({"OTHER": ["A"]})
    result = apply_operations(
        df, [{"operator": "codelist_check", "params": {"column": "SEX", "codelist": "SEX"}}], ct, {}
    )
    assert result["_pb_SEX_valid"].to_list() == [True]


def test_op_consistency_check_consistent(ct):
    df = _nw_df({"STUDYID": ["S001", "S001", "S001"]})
    result = apply_operations(
        df, [{"operator": "consistency_check", "params": {"column": "STUDYID"}}], ct, {}
    )
    assert all(result["_pb_STUDYID_consistent"].to_list())


def test_op_consistency_check_inconsistent(ct):
    df = _nw_df({"STUDYID": ["S001", "S001", "S002"]})
    result = apply_operations(
        df, [{"operator": "consistency_check", "params": {"column": "STUDYID"}}], ct, {}
    )
    vals = result["_pb_STUDYID_consistent"].to_list()
    assert vals[2] is False  # S002 is the outlier


def test_op_iso8601_check_valid(ct):
    df = _nw_df({"DTC": ["2024-01-01", "not-a-date", None]})
    result = apply_operations(
        df, [{"operator": "iso8601_check", "params": {"column": "DTC"}}], ct, {}
    )
    vals = result["_pb_DTC_iso8601"].to_list()
    assert vals == [True, False, True]


def test_op_column_presence_present(ct):
    df = _nw_df({"USUBJID": ["U1"]})
    result = apply_operations(
        df, [{"operator": "column_presence", "params": {"column": "USUBJID"}}], ct, {}
    )
    assert result["_pb_USUBJID_present"].to_list() == [True]


def test_op_column_presence_absent(ct):
    df = _nw_df({"OTHER": ["X"]})
    result = apply_operations(
        df, [{"operator": "column_presence", "params": {"column": "USUBJID"}}], ct, {}
    )
    assert result["_pb_USUBJID_present"].to_list() == [False]


def test_op_unknown_operator_skipped(ct):
    df = _nw_df({"x": [1]})
    # Should not raise; unknown operators are silently skipped.
    result = apply_operations(df, [{"operator": "nonexistent_op", "params": {}}], ct, {})
    assert result.columns == ["x"]


# ── NativeConformanceEngine: clean data ───────────────────────────────────────


def test_engine_clean_dm_all_pass(clean_result):
    assert clean_result.all_passed


def test_engine_clean_dm_zero_issues(clean_result):
    assert clean_result.n_total_issues == 0


def test_engine_rule_count(clean_result):
    assert len(clean_result.rule_results) == 426


def test_engine_result_types(clean_result):
    assert isinstance(clean_result, NativeConformanceResult)
    assert all(isinstance(r, NativeRuleResult) for r in clean_result.rule_results)


def test_engine_status_counts(clean_result):
    counts = clean_result.status_counts()
    assert STATUS_PASS in counts
    assert counts.get(STATUS_FAIL, 0) == 0


def test_engine_findings_empty_for_clean_data(clean_result):
    assert clean_result.findings() == []


# ── NativeConformanceEngine: violations ───────────────────────────────────────


def test_engine_detects_invalid_sex(engine):
    dm = pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["U1"],
            "SUBJID": ["1"],
            "SEX": ["Q"],
            "COUNTRY": ["USA"],
            "DMDTC": ["2024-01-01"],
        }
    )
    result = engine.run({"DM": dm})
    ids = {r.rule_id for r in result.rule_results if r.status == STATUS_FAIL}
    assert "SDTM-007" in ids


def test_engine_detects_invalid_country(engine):
    dm = pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["U1"],
            "SUBJID": ["1"],
            "COUNTRY": ["XYZ"],
            "DMDTC": ["2024-01-01"],
        }
    )
    result = engine.run({"DM": dm})
    ids = {r.rule_id for r in result.rule_results if r.status == STATUS_FAIL}
    assert "SDTM-009" in ids


def test_engine_detects_bad_date(engine):
    dm = pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["U1"],
            "SUBJID": ["1"],
            "DMDTC": ["not-a-date"],
        }
    )
    result = engine.run({"DM": dm})
    ids = {r.rule_id for r in result.rule_results if r.status == STATUS_FAIL}
    assert "SDTM-010" in ids


def test_engine_detects_null_usubjid(engine):
    dm = pl.DataFrame({"STUDYID": ["S1", "S1"], "DOMAIN": ["DM", "DM"], "USUBJID": ["U1", None]})
    result = engine.run({"DM": dm})
    ids = {r.rule_id for r in result.rule_results if r.status == STATUS_FAIL}
    assert "SDTM-004" in ids


def test_engine_detects_missing_dm_domain(engine):
    ae = pl.DataFrame({"STUDYID": ["S1"], "DOMAIN": ["AE"], "USUBJID": ["U1"]})
    result = engine.run({"AE": ae})
    ids = {r.rule_id for r in result.rule_results if r.status == STATUS_FAIL}
    assert "SDTM-001" in ids


def test_engine_detects_inconsistent_studyid(engine):
    dm = pl.DataFrame({"STUDYID": ["S1", "S2"], "DOMAIN": ["DM", "DM"], "USUBJID": ["U1", "U2"]})
    result = engine.run({"DM": dm})
    ids = {r.rule_id for r in result.rule_results if r.status == STATUS_FAIL}
    assert "SDTM-005" in ids


def test_engine_row_findings_populated(engine):
    dm = pl.DataFrame(
        {"STUDYID": ["S1"], "DOMAIN": ["DM"], "USUBJID": ["U1"], "SUBJID": ["1"], "SEX": ["Q"]}
    )
    result = engine.run({"DM": dm})
    findings = result.findings()
    assert len(findings) > 0
    f = findings[0]
    assert isinstance(f, NativeRowFinding)
    assert f.rule_id == "SDTM-007"
    assert f.row == 0


def test_engine_row_finding_has_usubjid(engine):
    dm = pl.DataFrame(
        {"STUDYID": ["S1"], "DOMAIN": ["DM"], "USUBJID": ["U1"], "SUBJID": ["1"], "SEX": ["Q"]}
    )
    result = engine.run({"DM": dm})
    findings = result.findings()
    sex_finding = next(f for f in findings if f.rule_id == "SDTM-007")
    assert sex_finding.usubjid == "U1"
    assert sex_finding.checked_column == "SEX"
    assert sex_finding.checked_value == "Q"


def test_engine_row_finding_has_date_column(engine):
    dm = pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["U1"],
            "SUBJID": ["1"],
            "SEX": ["M"],
            "RACE": ["WHITE"],
            "ETHNIC": ["NOT HISPANIC OR LATINO"],
            "COUNTRY": ["USA"],
            "ARMCD": ["A"],
            "ARM": ["Arm A"],
            "ACTARMCD": ["A"],
            "ACTARM": ["Arm A"],
            "DMDTC": ["not-a-date"],
        }
    )
    result = engine.run({"DM": dm})
    findings = result.findings()
    date_finding = next((f for f in findings if f.rule_id == "SDTM-010"), None)
    assert date_finding is not None
    assert date_finding.checked_column == "DMDTC"
    assert date_finding.checked_value == "not-a-date"
    assert date_finding.usubjid == "U1"


def test_engine_row_finding_no_usubjid_when_absent(engine):
    # Dataset without USUBJID
    dm = pl.DataFrame({"STUDYID": ["S1"], "DOMAIN": ["DM"], "SEX": ["Q"]})
    result = engine.run({"DM": dm})
    findings = result.findings()
    sex_finding = next((f for f in findings if f.rule_id == "SDTM-007"), None)
    if sex_finding is not None:
        assert sex_finding.usubjid is None


def test_engine_rules_status_filter(clean_result):
    passing = clean_result.rules(status=STATUS_PASS)
    assert all(r.status == STATUS_PASS for r in passing)


def test_engine_issues_list(engine):
    dm = pl.DataFrame(
        {"STUDYID": ["S1"], "DOMAIN": ["DM"], "USUBJID": ["U1"], "SUBJID": ["1"], "SEX": ["Q"]}
    )
    result = engine.run({"DM": dm})
    issues = result.issues()
    assert len(issues) > 0
    issue = next(i for i in issues if i["rule_id"] == "SDTM-007")
    assert issue["n_issues"] == 1
    assert issue["status"] == STATUS_FAIL


# ── Pandas backend ────────────────────────────────────────────────────────────


def test_engine_works_with_pandas(engine):
    dm = pd.DataFrame(
        {"STUDYID": ["S1"], "DOMAIN": ["DM"], "USUBJID": ["U1"], "SUBJID": ["1"], "SEX": ["Q"]}
    )
    result = engine.run({"DM": dm})
    ids = {r.rule_id for r in result.rule_results if r.status == STATUS_FAIL}
    assert "SDTM-007" in ids


# ── SubmissionPackage integration ─────────────────────────────────────────────


def test_submission_package_uses_rules_engine():
    pkg = pb.SubmissionPackage(
        datasets={"DM": _clean_dm()}, standard="sdtmig", standard_version="3.4"
    )
    report = pkg.validate_conformance()
    assert report.is_rules
    assert not report.is_core


def test_submission_package_summary_has_engine_key():
    pkg = pb.SubmissionPackage(datasets={"DM": _clean_dm()})
    report = pkg.validate_conformance()
    s = report.summary()
    assert s["engine"] == "built-in"
    assert "n_rules" in s
    assert "n_issues" in s


def test_submission_package_dirty_data_fails():
    dirty = pl.DataFrame(
        {"STUDYID": ["S1"], "DOMAIN": ["DM"], "USUBJID": ["U1"], "SUBJID": ["1"], "SEX": ["BAD"]}
    )
    pkg = pb.SubmissionPackage(datasets={"DM": dirty})
    report = pkg.validate_conformance()
    assert not report.all_passed()
    assert len(report.issues()) > 0


def test_submission_package_repr_shows_built_in_rules():
    pkg = pb.SubmissionPackage(datasets={"DM": _clean_dm()})
    report = pkg.validate_conformance()
    r = repr(report)
    assert "Built-in Rules" in r


def test_submission_package_to_json_rules(tmp_path):
    pkg = pb.SubmissionPackage(datasets={"DM": _clean_dm()})
    report = pkg.validate_conformance()
    dest = report.to_json(tmp_path / "r.json")
    data = json.loads(dest.read_text())
    assert data["summary"]["engine"] == "built-in"
    assert isinstance(data["issues"], list)


def test_submission_package_to_excel_rules(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl

    pkg = pb.SubmissionPackage(datasets={"DM": _clean_dm()})
    report = pkg.validate_conformance()
    dest = report.to_excel(tmp_path / "r.xlsx")
    wb = openpyxl.load_workbook(dest)
    assert "Rules_Report" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    # Every rule has a row.
    n_data_rows = wb["Rules_Report"].max_row - 1
    assert n_data_rows == len(report.rules())


def test_submission_package_rules_accessor():
    pkg = pb.SubmissionPackage(datasets={"DM": _clean_dm()})
    report = pkg.validate_conformance()
    rules = report.rules()
    assert len(rules) > 0
    assert all(isinstance(r, NativeRuleResult) for r in rules)


def test_submission_package_findings_accessor():
    dirty = pl.DataFrame({"STUDYID": ["S1"], "DOMAIN": ["DM"], "USUBJID": ["U1"], "SEX": ["BAD"]})
    pkg = pb.SubmissionPackage(datasets={"DM": dirty})
    report = pkg.validate_conformance()
    findings = report.findings()
    assert len(findings) > 0
    assert all(isinstance(f, NativeRowFinding) for f in findings)


def test_findings_df_returns_dataframe():
    import polars as pl

    dirty = pl.DataFrame({"STUDYID": ["S1"], "DOMAIN": ["DM"], "USUBJID": ["U1"], "SEX": ["BAD"]})
    report = pb.validate_sdtmig({"DM": dirty})
    df = report.findings_df()
    assert isinstance(df, pl.DataFrame)
    expected_cols = {
        "rule_id",
        "dataset",
        "row_index",
        "usubjid",
        "checked_column",
        "checked_value",
        "description",
    }
    assert expected_cols.issubset(set(df.columns))
    assert len(df) > 0


def test_findings_df_captures_correct_fields():
    import polars as pl

    dirty = pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["U99"],
            "SUBJID": ["99"],
            "SEX": ["Q"],
        }
    )
    report = pb.validate_sdtmig({"DM": dirty})
    df = report.findings_df()
    sex_row = df.filter(pl.col("rule_id") == "SDTM-007")
    assert len(sex_row) == 1
    assert sex_row["usubjid"][0] == "U99"
    assert sex_row["checked_column"][0] == "SEX"
    assert sex_row["checked_value"][0] == "Q"


def test_findings_df_empty_when_all_pass(clean_result):
    """findings_df() returns an empty DataFrame (correct schema) when no issues exist."""
    import polars as pl

    report = pb.ConformanceReport(native_result=clean_result)
    df = report.findings_df()
    assert isinstance(df, pl.DataFrame)
    assert len(df) == 0
    assert "rule_id" in df.columns


def test_findings_df_raises_for_core_report():
    """findings_df() raises TypeError on a CORE-backed report."""
    from pointblank.metadata._cdisc_core import ParsedCoreReport

    report = pb.ConformanceReport(core=ParsedCoreReport())
    with pytest.raises(TypeError, match="findings_df"):
        report.findings_df()


def test_get_findings_table_returns_gt():
    """get_findings_table() returns a GT object."""
    import polars as pl
    from great_tables import GT

    dirty = pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["U1"],
            "SUBJID": ["1"],
            "SEX": ["Q"],
        }
    )
    report = pb.validate_sdtmig({"DM": dirty})
    gt = report.get_findings_table()
    assert isinstance(gt, GT)
    html = gt._repr_html_()
    assert "SDTM-007" in html
    assert "U1" in html


def test_get_findings_table_raises_when_no_findings(clean_result):
    """get_findings_table() raises ValueError when there are no findings."""
    report = pb.ConformanceReport(native_result=clean_result)
    with pytest.raises(ValueError, match="No row-level findings"):
        report.get_findings_table()


# ── Phase 2: JSONata evaluator ────────────────────────────────────────────────

from pointblank.metadata._conformance.jsonata import (
    evaluate_jsonata,
    JSONataNotSupported,
    JSONataSyntaxError,
)


def test_jsonata_literals():
    assert evaluate_jsonata("true", {}) is True
    assert evaluate_jsonata("false", {}) is False
    assert evaluate_jsonata("null", {}) is None
    assert evaluate_jsonata("42", {}) == 42
    assert evaluate_jsonata("3.14", {}) == 3.14
    assert evaluate_jsonata('"hello"', {}) == "hello"


def test_jsonata_field_access():
    assert evaluate_jsonata("DOMAIN", {"DOMAIN": "AE"}) == "AE"
    assert evaluate_jsonata("MISSING", {"DOMAIN": "AE"}) is None


def test_jsonata_path_navigation():
    ctx = {"Dataset": {"Variable": "USUBJID"}}
    assert evaluate_jsonata("Dataset.Variable", ctx) == "USUBJID"


def test_jsonata_comparison():
    ctx = {"DOMAIN": "AE", "AGE": 30}
    assert evaluate_jsonata('DOMAIN = "AE"', ctx) is True
    assert evaluate_jsonata('DOMAIN != "DM"', ctx) is True
    assert evaluate_jsonata("AGE > 25", ctx) is True
    assert evaluate_jsonata("AGE < 25", ctx) is False
    assert evaluate_jsonata("AGE >= 30", ctx) is True
    assert evaluate_jsonata("AGE <= 30", ctx) is True


def test_jsonata_arithmetic():
    assert evaluate_jsonata("2 + 3", {}) == 5
    assert evaluate_jsonata("10 - 4", {}) == 6
    assert evaluate_jsonata("3 * 4", {}) == 12
    assert evaluate_jsonata("10 / 4", {}) == 2.5


def test_jsonata_boolean_operators():
    assert evaluate_jsonata("true and true", {}) is True
    assert evaluate_jsonata("true and false", {}) is False
    assert evaluate_jsonata("false or true", {}) is True
    assert evaluate_jsonata("not(false)", {}) is True
    assert evaluate_jsonata("not(true)", {}) is False


def test_jsonata_grouped():
    assert evaluate_jsonata("(2 + 3) * 4", {}) == 20


def test_jsonata_string_functions():
    assert evaluate_jsonata('$uppercase("ae")', {}) == "AE"
    assert evaluate_jsonata('$lowercase("AE")', {}) == "ae"
    assert evaluate_jsonata("$string(42)", {}) == "42"
    assert evaluate_jsonata('$length("hello")', {}) == 5
    assert evaluate_jsonata('$trim("  hi  ")', {}) == "hi"


def test_jsonata_substring():
    assert evaluate_jsonata('$substring("STUDYID", 0, 5)', {}) == "STUDY"
    assert evaluate_jsonata('$substring("STUDYID", 5)', {}) == "ID"
    assert evaluate_jsonata('$substring("hello", -3)', {}) == "llo"


def test_jsonata_aggregate_functions():
    assert evaluate_jsonata("$count(VALS)", {"VALS": [1, 2, 3]}) == 3
    assert evaluate_jsonata("$count(VALS)", {"VALS": None}) == 0
    assert evaluate_jsonata("$exists(X)", {"X": "y"}) is True
    assert evaluate_jsonata("$exists(X)", {"X": None}) is False
    assert evaluate_jsonata("$count($distinct(VALS))", {"VALS": [1, 2, 2, 3]}) == 3


def test_jsonata_context_field_expression():
    ctx = {"DOMAIN": "AE", "USUBJID": "S01"}
    assert evaluate_jsonata('$uppercase(DOMAIN) = "AE"', ctx) is True
    assert evaluate_jsonata("$length(DOMAIN) = 2", ctx) is True


def test_jsonata_not_supported_filter():
    import pytest

    # Filter expressions VALS[...] are not supported; raises either
    # JSONataNotSupported (when reached during evaluation) or JSONataSyntaxError
    # (when the parser hits unexpected '[' after consuming VALS).
    with pytest.raises((JSONataNotSupported, JSONataSyntaxError)):
        evaluate_jsonata("VALS[0]", {"VALS": [1, 2]})


def test_jsonata_syntax_error():
    import pytest

    with pytest.raises(JSONataSyntaxError):
        evaluate_jsonata("= broken", {})


def test_jsonata_empty_expression_returns_none():
    # covers the `if not tokens: return None` guard (line 466)
    assert evaluate_jsonata("", {}) is None
    assert evaluate_jsonata("   ", {}) is None


def test_jsonata_unrecognized_character_breaks_tokenizer():
    # `@` matches no token pattern; tokenizer breaks, returning []
    assert evaluate_jsonata("@", {}) is None


def test_jsonata_filter_expression_at_start_raises():
    # `[` as the START of a primary expression hits the not-supported guard
    with pytest.raises(JSONataNotSupported):
        evaluate_jsonata("[1, 2]", {})


def test_jsonata_unary_negative():
    assert evaluate_jsonata("-5", {}) == -5
    assert evaluate_jsonata("-X", {"X": 3}) == -3


def test_jsonata_not_missing_paren_raises():
    # `_expect("(")` fails when `not` is not followed by `(`
    with pytest.raises(JSONataSyntaxError):
        evaluate_jsonata("not true", {})


def test_jsonata_unexpected_end_of_expression():
    # `_primary` gets `t is None`
    with pytest.raises(JSONataSyntaxError):
        evaluate_jsonata("1 +", {})


def test_jsonata_bare_variable_reference():
    # bare `$x` (no parens) — parsed as ("var", "$x"), looked up in ctx
    assert evaluate_jsonata("$x", {"$x": 42}) == 42
    assert evaluate_jsonata("$y", {}) is None


def test_jsonata_path_trailing_dot():
    # trailing dot: parser breaks out of path loop when nothing follows the dot
    assert evaluate_jsonata("a.", {"a": 1}) == 1
    assert evaluate_jsonata("a.", {}) is None


def test_jsonata_path_non_identifier_after_dot():
    # path parsing breaks when a number follows the dot; parse fails
    with pytest.raises(JSONataSyntaxError):
        evaluate_jsonata("a.1", {})


def test_jsonata_path_non_dict_value_returns_none():
    # path traversal hits `else: value = None; break`
    assert evaluate_jsonata("a.b", {"a": "not_a_dict"}) is None
    assert evaluate_jsonata("a.b.c", {"a": 5}) is None


def test_jsonata_comparison_null_operand_returns_false():
    # ordering operators with a None side return False
    assert evaluate_jsonata("MISSING > 5", {}) is False
    assert evaluate_jsonata("MISSING < 5", {}) is False
    assert evaluate_jsonata("MISSING >= 5", {}) is False
    assert evaluate_jsonata("MISSING <= 5", {}) is False


def test_jsonata_binop_null_operand_returns_none():
    assert evaluate_jsonata("MISSING + 1", {}) is None
    assert evaluate_jsonata("MISSING - 1", {}) is None
    assert evaluate_jsonata("MISSING * 2", {}) is None
    assert evaluate_jsonata("MISSING / 2", {}) is None


def test_jsonata_length_null_and_list():
    assert evaluate_jsonata("$length(MISSING)", {}) == 0
    assert evaluate_jsonata("$length(VALS)", {"VALS": [1, 2, 3]}) == 3


def test_jsonata_count_scalar():
    assert evaluate_jsonata("$count(X)", {"X": "hello"}) == 1


def test_jsonata_distinct_null_and_scalar():
    assert evaluate_jsonata("$distinct(MISSING)", {}) == []
    assert evaluate_jsonata("$distinct(X)", {"X": "hello"}) == ["hello"]


def test_jsonata_sum():
    assert evaluate_jsonata("$sum(VALS)", {"VALS": [1, 2, 3]}) == 6
    assert evaluate_jsonata("$sum(X)", {"X": 5}) == 5
    assert evaluate_jsonata("$sum(X)", {"X": "str"}) == 0


def test_jsonata_max():
    assert evaluate_jsonata("$max(VALS)", {"VALS": [3, 1, 2]}) == 3
    assert evaluate_jsonata("$max(VALS)", {"VALS": []}) is None
    assert evaluate_jsonata("$max(X)", {"X": 7}) == 7


def test_jsonata_min():
    assert evaluate_jsonata("$min(VALS)", {"VALS": [3, 1, 2]}) == 1
    assert evaluate_jsonata("$min(VALS)", {"VALS": []}) is None
    assert evaluate_jsonata("$min(X)", {"X": 7}) == 7


def test_jsonata_round():
    assert evaluate_jsonata("$round(MISSING)", {}) is None
    assert evaluate_jsonata("$round(3.14, 1)", {}) == 3.1
    assert evaluate_jsonata("$round(3.14)", {}) == 3


def test_jsonata_floor_ceil_abs():
    assert evaluate_jsonata("$floor(3.7)", {}) == 3
    assert evaluate_jsonata("$floor(MISSING)", {}) is None
    assert evaluate_jsonata("$ceil(3.2)", {}) == 4
    assert evaluate_jsonata("$ceil(MISSING)", {}) is None
    assert evaluate_jsonata("$abs(-5)", {}) == 5
    assert evaluate_jsonata("$abs(MISSING)", {}) is None


def test_jsonata_type_function():
    assert evaluate_jsonata("$type(null)", {}) == "null"
    assert evaluate_jsonata("$type(true)", {}) == "boolean"
    assert evaluate_jsonata("$type(42)", {}) == "number"
    assert evaluate_jsonata('$type("hello")', {}) == "string"
    assert evaluate_jsonata("$type(VALS)", {"VALS": [1, 2]}) == "array"
    assert evaluate_jsonata("$type(D)", {"D": {"a": 1}}) == "object"
    # non-standard Python type not reachable through normal parsing
    from pointblank.metadata._conformance.jsonata import _call_function

    assert _call_function("type", [{1, 2}]) == "unknown"


def test_jsonata_not_function():
    assert evaluate_jsonata("$not(true)", {}) is False
    assert evaluate_jsonata("$not(false)", {}) is True


def test_jsonata_unsupported_function_raises():
    with pytest.raises(JSONataNotSupported):
        evaluate_jsonata("$unknown()", {})


# ── Phase 2: new operations ───────────────────────────────────────────────────


def test_has_required_variables_all_present():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    df = nw.from_native(
        pl.DataFrame({"STUDYID": ["X"], "DOMAIN": ["DM"], "USUBJID": ["U1"]}), eager_only=True
    )
    ct = ControlledTerminology({}, [])
    ops = [
        {
            "operator": "has_required_variables",
            "params": {"variables": ["STUDYID", "DOMAIN", "USUBJID"]},
        }
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_STUDYID_present"].to_list() == [True]
    assert result["_pb_DOMAIN_present"].to_list() == [True]
    assert result["_pb_USUBJID_present"].to_list() == [True]


def test_has_required_variables_missing():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    df = nw.from_native(pl.DataFrame({"STUDYID": ["X"]}), eager_only=True)
    ct = ControlledTerminology({}, [])
    ops = [{"operator": "has_required_variables", "params": {"variables": ["STUDYID", "DOMAIN"]}}]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_STUDYID_present"].to_list() == [True]
    assert result["_pb_DOMAIN_present"].to_list() == [False]


def test_valid_variable_order_correct():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    df = nw.from_native(
        pl.DataFrame({"STUDYID": ["X"], "DOMAIN": ["DM"], "USUBJID": ["U1"]}), eager_only=True
    )
    ct = ControlledTerminology({}, [])
    ops = [
        {
            "operator": "valid_variable_order",
            "params": {"expected_order": ["STUDYID", "DOMAIN", "USUBJID"]},
        }
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_variable_order_valid"].to_list() == [True]


def test_valid_variable_order_wrong():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    # DOMAIN appears before STUDYID
    df = nw.from_native(
        pl.DataFrame({"DOMAIN": ["DM"], "STUDYID": ["X"], "USUBJID": ["U1"]}), eager_only=True
    )
    ct = ControlledTerminology({}, [])
    ops = [
        {
            "operator": "valid_variable_order",
            "params": {"expected_order": ["STUDYID", "DOMAIN", "USUBJID"]},
        }
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_variable_order_valid"].to_list() == [False]


def test_valid_variable_order_absent_columns_skipped():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    # DOMAIN absent; remaining two are in order → True
    df = nw.from_native(pl.DataFrame({"STUDYID": ["X"], "USUBJID": ["U1"]}), eager_only=True)
    ct = ControlledTerminology({}, [])
    ops = [
        {
            "operator": "valid_variable_order",
            "params": {"expected_order": ["STUDYID", "DOMAIN", "USUBJID"]},
        }
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_variable_order_valid"].to_list() == [True]


def test_variable_type_check_numeric_ok():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    df = nw.from_native(pl.DataFrame({"AGE": [45.0]}), eager_only=True)
    ct = ControlledTerminology({}, [])
    ops = [
        {"operator": "variable_type_check", "params": {"column": "AGE", "expected_type": "numeric"}}
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_AGE_type_valid"].to_list() == [True]


def test_variable_type_check_numeric_fail():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    df = nw.from_native(pl.DataFrame({"AGE": ["45"]}), eager_only=True)  # string, not numeric
    ct = ControlledTerminology({}, [])
    ops = [
        {"operator": "variable_type_check", "params": {"column": "AGE", "expected_type": "numeric"}}
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_AGE_type_valid"].to_list() == [False]


def test_variable_type_check_absent_column_passes():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import polars as pl
    import narwhals as nw

    df = nw.from_native(pl.DataFrame({"STUDYID": ["X"]}), eager_only=True)
    ct = ControlledTerminology({}, [])
    ops = [
        {"operator": "variable_type_check", "params": {"column": "AGE", "expected_type": "numeric"}}
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_AGE_type_valid"].to_list() == [True]


# ── Phase 2: VARIABLE_METADATA_CHECK engine integration ──────────────────────


def _full_dm() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["S1-001"],
            "SUBJID": ["001"],
            "RFSTDTC": ["2020-01-01"],
            "RFENDTC": ["2020-06-30"],
            "SITEID": ["001"],
            "AGE": [45.0],
            "AGEU": ["YEARS"],
            "SEX": ["M"],
            "RACE": ["WHITE"],
            "ETHNIC": ["NOT HISPANIC OR LATINO"],
            "COUNTRY": ["USA"],
            "ARMCD": ["A"],
            "ARM": ["Arm A"],
            "ACTARMCD": ["A"],
            "ACTARM": ["Arm A"],
        }
    )


def test_variable_metadata_check_passes_for_complete_dm():
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["VARIABLE_METADATA_CHECK"])
    result = engine.run({"DM": _full_dm()})
    vmc = [r for r in result.rule_results if r.rule_type == "VARIABLE_METADATA_CHECK"]
    assert len(vmc) > 0
    # With a complete DM, all Fully Executable VMC rules on DM should pass.
    dm_rules = [
        r for r in vmc if "DM" in r.dataset and r.status not in ("not_supported", "not_applicable")
    ]
    assert all(r.status == "pass" for r in dm_rules), [
        (r.rule_id, r.status, r.message) for r in dm_rules
    ]


def test_variable_metadata_check_fails_missing_sex():
    dm = _full_dm().drop("SEX")
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["VARIABLE_METADATA_CHECK"])
    result = engine.run({"DM": dm})
    vmc = {r.rule_id: r for r in result.rule_results if r.rule_type == "VARIABLE_METADATA_CHECK"}
    # SDTM-032 checks SEX, RACE, ETHNIC, COUNTRY.
    assert vmc["SDTM-032"].status == "fail"


def test_variable_metadata_check_fails_wrong_order():
    dm = _full_dm().select(
        ["DOMAIN", "STUDYID"] + [c for c in _full_dm().columns if c not in ("DOMAIN", "STUDYID")]
    )
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["VARIABLE_METADATA_CHECK"])
    result = engine.run({"DM": dm})
    vmc = {r.rule_id: r for r in result.rule_results if r.rule_type == "VARIABLE_METADATA_CHECK"}
    assert vmc["SDTM-044"].status == "fail"


def test_partially_executable_returns_not_applicable():
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["VARIABLE_METADATA_CHECK"])
    result = engine.run({"DM": _full_dm()})
    # SDTM-049 and SDTM-050 require DEFINE dataset
    partial = {r.rule_id: r for r in result.rule_results if r.rule_id in ("SDTM-049", "SDTM-050")}
    assert partial["SDTM-049"].status == "not_applicable"
    assert partial["SDTM-050"].status == "not_applicable"
    assert "not provided" in (partial["SDTM-049"].message or "")


def test_partially_executable_runs_when_dataset_provided():
    # When the required dataset IS present, the rule should not return not_applicable.
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["VARIABLE_METADATA_CHECK"])
    # SDTM-049/050 have empty conditions so they'll pass on any dataset.
    stub_define = MetadataPackage(
        items={"DM": MetadataImport(source_format="cdisc_define", dataset_name="DM")}
    )
    result = engine.run({"DM": _full_dm()}, define_xml=stub_define)
    partial = {r.rule_id: r for r in result.rule_results if r.rule_id in ("SDTM-049", "SDTM-050")}
    assert all(r.status != "not_applicable" for r in partial.values())


# ── Phase 3: Define-XML operations and handlers ───────────────────────────────


def _make_var(
    name: str,
    dtype: str = "String",
    required: bool = False,
    allowed_values=None,
    display_format: str | None = None,
) -> VariableMetadata:
    return VariableMetadata(
        name=name,
        dtype=dtype,
        required=required,
        allowed_values=allowed_values,
        display_format=display_format,
    )


def _make_define_pkg(domain: str, variables: list[VariableMetadata]) -> MetadataPackage:
    meta = MetadataImport(
        source_format="cdisc_define", dataset_name=domain, domain=domain, variables=variables
    )
    return MetadataPackage(items={domain.upper(): meta})


def test_define_var_declared_present():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("STUDYID"), _make_var("SEX")],
    )
    df = nw.from_native(pl.DataFrame({"STUDYID": ["X"], "SEX": ["M"]}), eager_only=True)
    ct = ControlledTerminology({}, [])
    ops = [
        {"operator": "define_var_declared", "params": {"column": "STUDYID"}},
        {"operator": "define_var_declared", "params": {"column": "MISSING_VAR"}},
    ]
    result = apply_operations(df, ops, ct, {}, define_meta)
    assert result["_pb_STUDYID_in_define"].to_list() == [True]
    assert result["_pb_MISSING_VAR_in_define"].to_list() == [False]


def test_define_var_declared_no_define_meta():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    df = nw.from_native(pl.DataFrame({"STUDYID": ["X"]}), eager_only=True)
    ops = [{"operator": "define_var_declared", "params": {"column": "ANYTHING"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, None)
    assert result["_pb_ANYTHING_in_define"].to_list() == [True]


def test_define_required_check_passes_non_null():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("STUDYID", required=True)],
    )
    df = nw.from_native(pl.DataFrame({"STUDYID": ["S1", "S2"]}), eager_only=True)
    ops = [{"operator": "define_required_check", "params": {"column": "STUDYID"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_STUDYID_mandatory_ok"].to_list() == [True, True]


def test_define_required_check_flags_null():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("STUDYID", required=True)],
    )
    df = nw.from_native(pl.DataFrame({"STUDYID": ["S1", None]}), eager_only=True)
    ops = [{"operator": "define_required_check", "params": {"column": "STUDYID"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_STUDYID_mandatory_ok"].to_list() == [True, False]


def test_define_required_check_not_mandatory_always_true():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("OPTIONAL_VAR", required=False)],
    )
    df = nw.from_native(pl.DataFrame({"OPTIONAL_VAR": ["X", None]}), eager_only=True)
    ops = [{"operator": "define_required_check", "params": {"column": "OPTIONAL_VAR"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_OPTIONAL_VAR_mandatory_ok"].to_list() == [True, True]


def test_define_codelist_check_valid_values():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("SEX", allowed_values=["M", "F", "U"])],
    )
    df = nw.from_native(pl.DataFrame({"SEX": ["M", "F", "INVALID"]}), eager_only=True)
    ops = [{"operator": "define_codelist_check", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_SEX_define_valid"].to_list() == [True, True, False]


def test_define_codelist_check_null_passes():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("SEX", allowed_values=["M", "F"])],
    )
    df = nw.from_native(pl.DataFrame({"SEX": ["M", None]}), eager_only=True)
    ops = [{"operator": "define_codelist_check", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_SEX_define_valid"].to_list() == [True, True]


def test_define_codelist_check_no_codelist_always_true():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("NOTES")],  # no allowed_values
    )
    df = nw.from_native(pl.DataFrame({"NOTES": ["anything"]}), eager_only=True)
    ops = [{"operator": "define_codelist_check", "params": {"column": "NOTES"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_NOTES_define_valid"].to_list() == [True]


def test_define_type_check_numeric_ok():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("AGE", dtype="Float64", display_format="float")],
    )
    df = nw.from_native(pl.DataFrame({"AGE": [45.0]}), eager_only=True)
    ops = [{"operator": "define_type_check", "params": {"column": "AGE"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_AGE_define_type_ok"].to_list() == [True]


def test_define_type_check_char_mismatch():
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance.ct import ControlledTerminology
    import narwhals as nw

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[_make_var("STUDYID", display_format="text")],
    )
    df = nw.from_native(pl.DataFrame({"STUDYID": [1, 2]}), eager_only=True)  # numeric, not text
    ops = [{"operator": "define_type_check", "params": {"column": "STUDYID"}}]
    result = apply_operations(df, ops, ControlledTerminology({}, []), {}, define_meta)
    assert result["_pb_STUDYID_define_type_ok"].to_list() == [False, False]


# ── Phase 3: engine integration ───────────────────────────────────────────────


def _dm_with_bad_sex() -> pl.DataFrame:
    return pl.DataFrame(
        {
            "STUDYID": ["S1"],
            "DOMAIN": ["DM"],
            "USUBJID": ["S1-001"],
            "SUBJID": ["001"],
            "SEX": ["INVALID"],
            "RACE": ["WHITE"],
            "ETHNIC": ["NOT HISPANIC OR LATINO"],
            "COUNTRY": ["USA"],
            "AGE": [45.0],
            "AGEU": ["YEARS"],
            "SITEID": ["001"],
            "RFSTDTC": ["2020-01-01"],
            "RFENDTC": ["2020-06-30"],
            "ARMCD": ["A"],
            "ARM": ["Arm A"],
            "ACTARMCD": ["A"],
            "ACTARM": ["Arm A"],
        }
    )


def _dm_define_pkg() -> MetadataPackage:
    return _make_define_pkg(
        "DM",
        [
            _make_var("STUDYID", required=True),
            _make_var("DOMAIN", required=True),
            _make_var("USUBJID", required=True),
            _make_var("SEX", allowed_values=["M", "F", "U", "UNDIFFERENTIATED"]),
            _make_var("RACE", allowed_values=["WHITE", "BLACK OR AFRICAN AMERICAN", "ASIAN"]),
            _make_var("ETHNIC", allowed_values=["NOT HISPANIC OR LATINO", "HISPANIC OR LATINO"]),
            _make_var("AGE", dtype="Float64", display_format="float"),
        ],
    )


def test_define_item_metadata_check_all_declared():
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DEFINE_ITEM_METADATA_CHECK"])
    pkg = _dm_define_pkg()
    result = engine.run({"DM": _clean_dm()}, define_xml=pkg)
    sdtm_051 = next(r for r in result.rule_results if r.rule_id == "SDTM-051")
    assert sdtm_051.status == "pass"


def test_define_item_metadata_check_undeclared_variable():
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DEFINE_ITEM_METADATA_CHECK"])
    # Provide a Define-XML that does NOT declare DOMAIN
    pkg = _make_define_pkg(
        "DM",
        [
            _make_var("STUDYID", required=True),
            _make_var("USUBJID"),
            _make_var("SEX"),
            _make_var("AGE", display_format="float"),
        ],
    )
    result = engine.run({"DM": _clean_dm()}, define_xml=pkg)
    sdtm_051 = next(r for r in result.rule_results if r.rule_id == "SDTM-051")
    assert sdtm_051.status == "fail"


def test_define_codelist_check_flags_bad_value():
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DEFINE_CODELIST_CHECK"])
    pkg = _dm_define_pkg()
    result = engine.run({"DM": _dm_with_bad_sex()}, define_xml=pkg)
    sdtm_056 = next(r for r in result.rule_results if r.rule_id == "SDTM-056")
    assert sdtm_056.status == "fail"
    assert sdtm_056.n_issues == 1


def test_define_codelist_check_passes_valid_values():
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DEFINE_CODELIST_CHECK"])
    pkg = _dm_define_pkg()
    result = engine.run({"DM": _clean_dm()}, define_xml=pkg)
    sdtm_056 = next(r for r in result.rule_results if r.rule_id == "SDTM-056")
    assert sdtm_056.status == "pass"


def test_define_rules_not_applicable_without_define_xml():
    engine = NativeConformanceEngine(
        "sdtmig",
        "3.4",
        rule_types=["DEFINE_ITEM_METADATA_CHECK", "DEFINE_CODELIST_CHECK"],
    )
    result = engine.run({"DM": _clean_dm()})  # no define_xml
    for r in result.rule_results:
        assert r.status == "not_applicable", f"{r.rule_id} was {r.status}"


def test_define_rules_applicable_with_define_xml():
    engine = NativeConformanceEngine(
        "sdtmig",
        "3.4",
        rule_types=["DEFINE_ITEM_METADATA_CHECK", "DEFINE_CODELIST_CHECK"],
    )
    result = engine.run({"DM": _clean_dm()}, define_xml=_dm_define_pkg())
    statuses = {r.status for r in result.rule_results}
    assert "not_applicable" not in statuses


def test_engine_accepts_metadata_import_directly():
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DEFINE_ITEM_METADATA_CHECK"])
    dm_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[
            _make_var("STUDYID", required=True),
            _make_var("DOMAIN"),
            _make_var("USUBJID"),
            _make_var("SEX"),
            _make_var("AGE", display_format="float"),
        ],
    )
    result = engine.run({"DM": _clean_dm()}, define_xml=dm_meta)
    sdtm_051 = next(r for r in result.rule_results if r.rule_id == "SDTM-051")
    assert sdtm_051.status in ("pass", "fail")  # executed, not not_applicable


def test_engine_rule_count_phase3():
    engine = NativeConformanceEngine("sdtmig", "3.4")
    result = engine.run({"DM": _clean_dm()})
    assert len(result.rule_results) == 426


_FIXTURES_DIR = Path(__file__).parent / "metadata_fixtures"


def test_load_define_xml_from_path_returns_package():
    pytest.importorskip("lxml")
    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DEFINE_ITEM_METADATA_CHECK"])
    # Pass a filesystem path; _load_define_xml reads it and returns a MetadataPackage
    result = engine.run({"DM": _clean_dm()}, define_xml=_FIXTURES_DIR / "define.xml")
    assert result is not None
    define_rules = [r for r in result.rule_results if r.rule_type == "DEFINE_ITEM_METADATA_CHECK"]
    assert any(r.status != "not_applicable" for r in define_rules)


def test_load_define_xml_from_path_wraps_single_import(monkeypatch):
    pytest.importorskip("lxml")
    from pointblank.metadata._types import MetadataImport

    single = MetadataImport(source_format="cdisc_define", dataset_name="DM", domain="DM")
    monkeypatch.setattr(
        "pointblank.metadata._conformance.engine._read_define_xml_metadata",
        lambda path: single,
        raising=False,
    )
    # Patch the import inside the try block so the monkeypatched name is used
    import pointblank.metadata._conformance.engine as eng_mod
    import pointblank.metadata._readers_cdisc as readers_mod

    monkeypatch.setattr(readers_mod, "_read_define_xml_metadata", lambda path: single)

    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DEFINE_ITEM_METADATA_CHECK"])
    # Trigger _load_define_xml with a path string; it will call the patched reader
    pkg = eng_mod.NativeConformanceEngine._load_define_xml(
        engine, str(_FIXTURES_DIR / "define.xml")
    )
    assert pkg is not None
    assert "DM" in pkg.items


def test_load_define_xml_import_error_returns_none(monkeypatch):
    import builtins
    import pointblank.metadata._conformance.engine as eng_mod

    real_import = builtins.__import__

    def mock_import(name, *args, **kwargs):
        if name == "pointblank.metadata._readers_cdisc":
            raise ImportError("simulated missing lxml")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", mock_import)
    engine = NativeConformanceEngine("sdtmig", "3.4")
    pkg = eng_mod.NativeConformanceEngine._load_define_xml(engine, "/some/path/define.xml")
    assert pkg is None


def test_ct_repr():
    """ControlledTerminology.__repr__ returns expected string."""
    from pointblank.metadata._conformance.ct import ControlledTerminology

    ct = ControlledTerminology({"SEX": {"M", "F"}}, ["2023-12-01"])
    r = repr(ct)
    assert "ControlledTerminology" in r
    assert "packages=" in r
    assert "n_codelists=" in r


def test_ct_available_no_dir(tmp_path, monkeypatch):
    """ControlledTerminology.available() returns [] when _CT_DIR doesn't exist."""
    from pointblank.metadata._conformance import ct as ct_module

    monkeypatch.setattr(ct_module, "_CT_DIR", tmp_path / "nonexistent")
    result = ct_module.ControlledTerminology.available()
    assert result == []


def test_ct_load_default_empty(tmp_path, monkeypatch):
    """ControlledTerminology.load_default() returns empty CT when no packages available."""
    from pointblank.metadata._conformance import ct as ct_module

    empty_dir = tmp_path / "ct_empty"
    empty_dir.mkdir()
    monkeypatch.setattr(ct_module, "_CT_DIR", empty_dir)
    ct = ct_module.ControlledTerminology.load_default()
    assert isinstance(ct, ct_module.ControlledTerminology)
    assert ct.packages == []


def test_native_rule_applies_to_domain_match():
    from pointblank.metadata._conformance.rule_loader import NativeRule

    rule = NativeRule(
        core_id="X-001",
        rule_type="RECORD_CHECK",
        executability="Fully Executable",
        sensitivity="Error",
        description="test",
        authority="CDISC",
        domains=["DM", "AE"],
        standards=["sdtmig"],
        classes=[],
        operations=[],
        conditions={},
        actions={},
    )
    assert rule.applies_to_domain("DM") is True
    assert rule.applies_to_domain("XX") is False


def test_native_rule_applies_to_standard_match():
    from pointblank.metadata._conformance.rule_loader import NativeRule

    rule = NativeRule(
        core_id="X-001",
        rule_type="RECORD_CHECK",
        executability="Fully Executable",
        sensitivity="Error",
        description="test",
        authority="CDISC",
        domains=[],
        standards=["sdtmig"],
        classes=[],
        operations=[],
        conditions={},
        actions={},
    )
    assert rule.applies_to_standard("sdtmig") is True
    assert rule.applies_to_standard("adamig") is False


def test_rule_loader_available_no_dir(tmp_path, monkeypatch):
    from pointblank.metadata._conformance import rule_loader as rl_module

    monkeypatch.setattr(rl_module, "_DATA_DIR", tmp_path / "nonexistent")
    result = rl_module.RuleLoader.available()
    assert result == []


def test_rule_loader_available_invalid_json(tmp_path, monkeypatch):
    from pointblank.metadata._conformance import rule_loader as rl_module

    monkeypatch.setattr(rl_module, "_DATA_DIR", tmp_path)
    (tmp_path / "bad.json").write_text("not valid json", encoding="utf-8")
    result = rl_module.RuleLoader.available()
    assert result == []


def test_rule_loader_catalog_metadata_missing_raises():
    from pointblank.metadata._conformance.rule_loader import RuleLoader

    with pytest.raises(FileNotFoundError):
        RuleLoader.catalog_metadata("imaginary", "99.0")


def test_apply_operations_handler_exception(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": ["M"]}))
    ops = [{"operator": "codelist_check", "params": {"column": "SEX", "codelist": "SEX"}}]
    result = apply_operations(df, ops, ct, {})
    assert "_pb_SEX_valid" in result.columns


def test_apply_operations_exception_silenced(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._conformance import operations as ops_module
    import narwhals as nw
    import polars as pl

    def _raise(df, params, ct, datasets, define_meta=None):
        raise RuntimeError("simulated failure")

    original_registry = ops_module._REGISTRY.copy()
    ops_module._REGISTRY["codelist_check"] = _raise
    try:
        df = nw.from_native(pl.DataFrame({"SEX": ["M"]}))
        ops = [{"operator": "codelist_check", "params": {"column": "SEX", "codelist": "SEX"}}]
        result = apply_operations(df, ops, ct, {})
        assert "_pb_SEX_valid" not in result.columns
    finally:
        ops_module._REGISTRY["codelist_check"] = original_registry["codelist_check"]


def test_op_codelist_check_unknown_codelist(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": ["M"]}))
    ops = [
        {"operator": "codelist_check", "params": {"column": "SEX", "codelist": "NONEXISTENT_CL"}}
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_SEX_valid"].to_list() == [True]


def test_op_consistency_check_missing_column(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"OTHER": [1, 2, 3]}))
    ops = [{"operator": "consistency_check", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_SEX_consistent"].to_list() == [True, True, True]


def test_op_consistency_check_all_null(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": pl.Series([None, None, None], dtype=pl.String)}))
    ops = [{"operator": "consistency_check", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_SEX_consistent"].to_list() == [True, True, True]


def test_op_unique_per_subject_no_usubjid(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": ["M", "F"]}))
    ops = [{"operator": "unique_per_subject", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_SEX_unique"].to_list() == [True, True]


def test_op_unique_per_subject_duplicates(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(
        pl.DataFrame({"USUBJID": ["S1", "S1", "S2"], "AETERM": ["headache", "headache", "fever"]})
    )
    ops = [{"operator": "unique_per_subject", "params": {"column": "AETERM"}}]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_AETERM_unique"].to_list() == [False, False, True]


def test_op_variable_type_check_numeric(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"AGE": [30, 40, 50]}))
    ops = [
        {"operator": "variable_type_check", "params": {"column": "AGE", "expected_type": "numeric"}}
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_AGE_type_valid"].to_list() == [True, True, True]


def test_op_variable_type_check_character(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": ["M", "F"]}))
    ops = [
        {
            "operator": "variable_type_check",
            "params": {"column": "SEX", "expected_type": "character"},
        }
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_SEX_type_valid"].to_list() == [True, True]


def test_op_variable_type_check_unknown_type(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": ["M", "F"]}))
    ops = [
        {
            "operator": "variable_type_check",
            "params": {"column": "SEX", "expected_type": "unknown_category"},
        }
    ]
    result = apply_operations(df, ops, ct, {})
    assert result["_pb_SEX_type_valid"].to_list() == [True, True]


def test_op_define_required_check_no_define_meta(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"STUDYID": ["S1", "S2"]}))
    ops = [{"operator": "define_required_check", "params": {"column": "STUDYID"}}]
    result = apply_operations(df, ops, ct, {}, define_meta=None)
    assert result["_pb_STUDYID_mandatory_ok"].to_list() == [True, True]


def test_op_define_codelist_check_no_define_meta(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": ["M", "F"]}))
    ops = [{"operator": "define_codelist_check", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ct, {}, define_meta=None)
    assert result["_pb_SEX_define_valid"].to_list() == [True, True]


def test_op_define_type_check_no_define_meta(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    import narwhals as nw
    import polars as pl

    df = nw.from_native(pl.DataFrame({"SEX": ["M", "F"]}))
    ops = [{"operator": "define_type_check", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ct, {}, define_meta=None)
    assert result["_pb_SEX_define_type_ok"].to_list() == [True, True]


def test_op_define_type_check_var_meta_none(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._types import MetadataImport, VariableMetadata
    import narwhals as nw
    import polars as pl

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[VariableMetadata(name="STUDYID")],
    )
    df = nw.from_native(pl.DataFrame({"AGE": [30, 40]}))
    ops = [{"operator": "define_type_check", "params": {"column": "AGE"}}]
    result = apply_operations(df, ops, ct, {}, define_meta=define_meta)
    assert result["_pb_AGE_define_type_ok"].to_list() == [True, True]


def test_op_define_type_check_unknown_define_type(ct):
    from pointblank.metadata._conformance.operations import apply_operations
    from pointblank.metadata._types import MetadataImport, VariableMetadata
    import narwhals as nw
    import polars as pl

    define_meta = MetadataImport(
        source_format="cdisc_define",
        dataset_name="DM",
        domain="DM",
        variables=[VariableMetadata(name="SEX", display_format="custom_unknown_type")],
    )
    df = nw.from_native(pl.DataFrame({"SEX": ["M", "F"]}))
    ops = [{"operator": "define_type_check", "params": {"column": "SEX"}}]
    result = apply_operations(df, ops, ct, {}, define_meta=define_meta)
    assert result["_pb_SEX_define_type_ok"].to_list() == [True, True]


def test_native_rule_applies_to_domain_empty():
    from pointblank.metadata._conformance.rule_loader import NativeRule

    rule = NativeRule(
        core_id="X-001",
        rule_type="RECORD_CHECK",
        executability="Fully Executable",
        sensitivity="Error",
        description="test",
        authority="CDISC",
        domains=[],
        standards=["sdtmig"],
        classes=[],
        operations=[],
        conditions={},
        actions={},
    )
    assert rule.applies_to_domain("DM") is True


def test_native_rule_applies_to_standard_empty():
    from pointblank.metadata._conformance.rule_loader import NativeRule

    rule = NativeRule(
        core_id="X-001",
        rule_type="RECORD_CHECK",
        executability="Fully Executable",
        sensitivity="Error",
        description="test",
        authority="CDISC",
        domains=["DM"],
        standards=[],
        classes=[],
        operations=[],
        conditions={},
        actions={},
    )
    assert rule.applies_to_standard("any") is True


def test_engine_with_explicit_ct_packages():
    from pointblank.metadata._conformance.ct import ControlledTerminology

    available = ControlledTerminology.available()
    if not available:
        pytest.skip("No CT packages available")
    engine = NativeConformanceEngine("sdtmig", "3.4", ct_packages=available)
    assert engine.ct_packages == available


def test_engine_evaluate_unsupported_rule_type():
    from pointblank.metadata._conformance.engine import NativeConformanceEngine
    from pointblank.metadata._conformance.rule_loader import NativeRule
    from pointblank.metadata._conformance.result import STATUS_NOT_SUPPORTED

    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["RECORD_CHECK"])
    rule = NativeRule(
        core_id="FAKE-001",
        rule_type="UNSUPPORTED_RULE_TYPE",
        executability="Fully Executable",
        sensitivity="Error",
        description="fake",
        authority="TEST",
        domains=[],
        standards=[],
        classes=[],
        operations=[],
        conditions={},
        actions={},
    )
    result = engine._evaluate_rule(rule, {})
    assert result.status == STATUS_NOT_SUPPORTED


def test_engine_partially_executable_missing_dataset():
    from pointblank.metadata._conformance.engine import NativeConformanceEngine
    from pointblank.metadata._conformance.rule_loader import NativeRule
    from pointblank.metadata._conformance.result import STATUS_NOT_APPLICABLE

    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["RECORD_CHECK"])
    rule = NativeRule(
        core_id="PART-001",
        rule_type="RECORD_CHECK",
        executability="Partially Executable",
        sensitivity="Error",
        description="needs EX dataset",
        authority="TEST",
        domains=["DM"],
        standards=[],
        classes=[],
        datasets=["EX"],
        operations=[],
        conditions={},
        actions={},
    )
    result = engine._evaluate_rule(rule, {"DM": nw.from_native(pl.DataFrame({"STUDYID": ["S1"]}))})
    assert result.status == STATUS_NOT_APPLICABLE


def test_engine_evaluate_rule_exception_returns_error_status():
    from pointblank.metadata._conformance.engine import NativeConformanceEngine
    from pointblank.metadata._conformance.rule_loader import NativeRule
    from pointblank.metadata._conformance.result import STATUS_ERROR

    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["RECORD_CHECK"])
    rule = NativeRule(
        core_id="ERR-001",
        rule_type="RECORD_CHECK",
        executability="Fully Executable",
        sensitivity="Error",
        description="error rule",
        authority="TEST",
        domains=["DM"],
        standards=[],
        classes=[],
        operations=[],
        conditions={},
        actions={},
    )
    original = engine._record_check

    def _raising(r, ds):
        raise RuntimeError("simulated failure")

    engine._record_check = _raising
    try:
        result = engine._evaluate_rule(rule, {"DM": nw.from_native(pl.DataFrame({"X": [1]}))})
        assert result.status == STATUS_ERROR
    finally:
        engine._record_check = original


def test_engine_dataset_metadata_check_evaluation_error():
    from pointblank.metadata._conformance.engine import NativeConformanceEngine
    from pointblank.metadata._conformance.rule_loader import NativeRule
    from pointblank.metadata._conformance.result import STATUS_PASS

    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DATASET_METADATA_CHECK"])
    rule = NativeRule(
        core_id="EVAL-001",
        rule_type="DATASET_METADATA_CHECK",
        executability="Fully Executable",
        sensitivity="Error",
        description="bad conditions",
        authority="TEST",
        domains=["DM"],
        standards=[],
        classes=[],
        operations=[],
        conditions={
            "all": [{"name": "_pb_NONEXISTENT_col", "operator": "equal_to", "value": True}]
        },
        actions={},
    )
    result = engine._evaluate_rule(rule, {"DM": nw.from_native(pl.DataFrame({"X": [1]}))})
    assert result.status == STATUS_PASS


def test_engine_domain_presence_check_prohibited_present():
    from pointblank.metadata._conformance.engine import NativeConformanceEngine
    from pointblank.metadata._conformance.rule_loader import NativeRule
    from pointblank.metadata._conformance.result import STATUS_FAIL

    engine = NativeConformanceEngine("sdtmig", "3.4", rule_types=["DOMAIN_PRESENCE_CHECK"])
    rule = NativeRule(
        core_id="PRES-001",
        rule_type="DOMAIN_PRESENCE_CHECK",
        executability="Fully Executable",
        sensitivity="Error",
        description="DM should not be present",
        authority="TEST",
        domains=[],
        standards=[],
        classes=[],
        operations=[],
        conditions={},
        actions={"params": {"required_domains": [], "prohibited_domains": ["DM"]}},
    )
    result = engine._evaluate_rule(rule, {"DM": nw.from_native(pl.DataFrame({"X": [1]}))})
    assert result.status == STATUS_FAIL
    assert "Prohibited" in result.message


# ── CDISC CORE parser (_cdisc_core.py) ────────────────────────────────────────


def _minimal_report(**extra):
    base = {
        "Conformance_Details": {
            "Standard": "SDTMIG",
            "Version": "V3.4",
            "CORE_Engine_Version": "0.16.0",
        },
        "Dataset_Details": [{"filename": "dm.xpt"}],
        "Issue_Summary": [{"dataset": "DM", "core_id": "CORE-001", "message": "msg", "issues": 2}],
        "Issue_Details": [
            {
                "core_id": "CORE-001",
                "message": "issue msg",
                "dataset": "DM",
                "executability": "fully executable",
                "USUBJID": "",
                "row": "1",
                "SEQ": "",
                "variables": ["VAR"],
                "values": ["bad"],
            }
        ],
        "Rules_Report": [
            {
                "core_id": "CORE-001",
                "status": "ISSUE REPORTED",
                "message": "m",
                "version": "1",
                "cdisc_rule_id": "CG01",
                "fda_rule_id": "",
            },
            {
                "core_id": "CORE-002",
                "status": "SUCCESS",
                "message": "",
                "version": "1",
                "cdisc_rule_id": "",
                "fda_rule_id": "",
            },
            {
                "core_id": "CORE-003",
                "status": "SKIPPED",
                "message": "",
                "version": "1",
                "cdisc_rule_id": "",
                "fda_rule_id": "",
            },
            {
                "core_id": "CORE-004",
                "status": "EXECUTION ERROR",
                "message": "err",
                "version": "1",
                "cdisc_rule_id": "",
                "fda_rule_id": "",
            },
        ],
    }
    base.update(extra)
    return base


def test_cdisc_core_parse_minimal_report():
    from pointblank.metadata._cdisc_core import parse_core_report, ParsedCoreReport

    parsed = parse_core_report(_minimal_report())
    assert isinstance(parsed, ParsedCoreReport)
    assert parsed.standard == "SDTMIG"
    assert parsed.version == "V3.4"
    assert parsed.engine_version == "0.16.0"


def test_cdisc_core_n_total_issues():
    from pointblank.metadata._cdisc_core import parse_core_report

    parsed = parse_core_report(_minimal_report())
    assert parsed.n_total_issues == 2


def test_cdisc_core_status_counts():
    from pointblank.metadata._cdisc_core import (
        parse_core_report,
        STATUS_SUCCESS,
        STATUS_SKIPPED,
        STATUS_ISSUE,
        STATUS_ERROR,
    )

    parsed = parse_core_report(_minimal_report())
    counts = parsed.status_counts()
    assert counts[STATUS_ISSUE] == 1
    assert counts[STATUS_SUCCESS] == 1
    assert counts[STATUS_SKIPPED] == 1
    assert counts[STATUS_ERROR] == 1


def test_cdisc_core_failing_rules():
    from pointblank.metadata._cdisc_core import parse_core_report

    parsed = parse_core_report(_minimal_report())
    failing = parsed.failing_rules()
    assert len(failing) == 2
    assert all(r.is_failing for r in failing)


def test_cdisc_core_all_passed_false():
    from pointblank.metadata._cdisc_core import parse_core_report

    parsed = parse_core_report(_minimal_report())
    assert parsed.all_passed is False


def test_cdisc_core_all_passed_true():
    from pointblank.metadata._cdisc_core import parse_core_report

    report = {
        "Conformance_Details": {},
        "Issue_Summary": [],
        "Issue_Details": [],
        "Rules_Report": [{"core_id": "CORE-1", "status": "SUCCESS"}],
    }
    parsed = parse_core_report(report)
    assert parsed.all_passed is True


def test_cdisc_core_all_passed_no_rules_fallback():
    from pointblank.metadata._cdisc_core import parse_core_report

    report = {
        "Conformance_Details": {},
        "Issue_Summary": [],
        "Issue_Details": [],
        "Rules_Report": [],
    }
    parsed = parse_core_report(report)
    assert parsed.all_passed is True


def test_cdisc_core_finding_fields():
    from pointblank.metadata._cdisc_core import parse_core_report, CoreFinding

    parsed = parse_core_report(_minimal_report())
    assert len(parsed.findings) == 1
    f = parsed.findings[0]
    assert isinstance(f, CoreFinding)
    assert f.rule_id == "CORE-001"
    assert f.dataset == "DM"
    assert f.row == 1
    assert f.usubjid is None
    assert f.seq is None
    assert f.variables == ["VAR"]
    assert f.values == ["bad"]
    assert f.executability == "fully executable"


def test_cdisc_core_rule_result_is_failing_property():
    from pointblank.metadata._cdisc_core import (
        CoreRuleResult,
        STATUS_SUCCESS,
        STATUS_ISSUE,
        STATUS_ERROR,
        STATUS_SKIPPED,
    )

    assert CoreRuleResult(rule_id="C", status=STATUS_SUCCESS).is_failing is False
    assert CoreRuleResult(rule_id="C", status=STATUS_SKIPPED).is_failing is False
    assert CoreRuleResult(rule_id="C", status=STATUS_ISSUE).is_failing is True
    assert CoreRuleResult(rule_id="C", status=STATUS_ERROR).is_failing is True


def test_cdisc_core_issue_summary_parsed():
    from pointblank.metadata._cdisc_core import parse_core_report, CoreIssueSummary

    parsed = parse_core_report(_minimal_report())
    assert len(parsed.issue_summary) == 1
    s = parsed.issue_summary[0]
    assert isinstance(s, CoreIssueSummary)
    assert s.dataset == "DM"
    assert s.rule_id == "CORE-001"
    assert s.issues == 2


def test_cdisc_core_datasets_parsed():
    from pointblank.metadata._cdisc_core import parse_core_report

    parsed = parse_core_report(_minimal_report())
    assert len(parsed.datasets) == 1
    assert parsed.datasets[0]["filename"] == "dm.xpt"


def test_cdisc_core_as_int():
    from pointblank.metadata._cdisc_core import _as_int

    assert _as_int(None) is None
    assert _as_int("") is None
    assert _as_int("5") == 5
    assert _as_int(3) == 3
    assert _as_int("bad") is None


def test_cdisc_core_empty_to_none():
    from pointblank.metadata._cdisc_core import _empty_to_none

    assert _empty_to_none("") is None
    assert _empty_to_none("hello") == "hello"
    assert _empty_to_none(None) is None


def test_cdisc_core_normalize_version():
    from pointblank.metadata._cdisc_core import _normalize_version

    assert _normalize_version("3.4") == "3-4"
    assert _normalize_version("1.1.1") == "1-1-1"
    assert _normalize_version("3-4") == "3-4"


def test_cdisc_core_resolve_core_command_str():
    from pointblank.metadata._cdisc_core import _resolve_core_command

    assert _resolve_core_command("core") == ["core"]


def test_cdisc_core_resolve_core_command_sequence():
    from pointblank.metadata._cdisc_core import _resolve_core_command

    assert _resolve_core_command(["python", "core.py"]) == ["python", "core.py"]


def test_cdisc_core_resolve_core_command_env(monkeypatch):
    from pointblank.metadata._cdisc_core import _resolve_core_command

    monkeypatch.setenv("POINTBLANK_CDISC_CORE", "python /path/to/core.py")
    result = _resolve_core_command(None)
    assert result == ["python", "/path/to/core.py"]


def test_cdisc_core_resolve_core_command_not_found(monkeypatch):
    from pointblank.metadata._cdisc_core import _resolve_core_command, CoreNotFoundError
    import shutil

    monkeypatch.delenv("POINTBLANK_CDISC_CORE", raising=False)
    monkeypatch.setattr("pointblank.metadata._cdisc_core.shutil.which", lambda name: None)
    with pytest.raises(CoreNotFoundError):
        _resolve_core_command(None)


def test_cdisc_core_runner_properties():
    from pointblank.metadata._cdisc_core import _CoreRunner

    runner = _CoreRunner(core="mycore", cwd="/tmp")
    assert runner.command == ["mycore"]
    assert runner.cwd == "/tmp"


def test_cdisc_core_runner_cwd_none():
    from pointblank.metadata._cdisc_core import _CoreRunner

    runner = _CoreRunner(core="mycore")
    assert runner.cwd is None


def test_cdisc_core_parse_raises_on_non_dict():
    from pointblank.metadata._cdisc_core import parse_core_report

    with pytest.raises(TypeError):
        parse_core_report(["not", "a", "dict"])


def test_cdisc_core_parse_raises_on_unrecognized_keys():
    from pointblank.metadata._cdisc_core import parse_core_report

    with pytest.raises(ValueError):
        parse_core_report({"foo": 1, "bar": 2})
